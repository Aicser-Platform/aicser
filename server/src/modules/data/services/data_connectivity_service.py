"""
Universal Data Connectivity Service
Handles file uploads, database connections, and data source management
"""

import asyncio
import io
import logging
import os
import pandas as pd
import json
import re
import sqlalchemy as sa
from typing import Dict, List, Optional, Any, Union
from datetime import datetime, date
import tempfile
import time
from pathlib import Path
from .database_connector_service import DatabaseConnectorService
from .upload_datasource_storage_service import UploadDatasourceStorageService
try:
    from ee.modules.data.services.ai_schema_service import AISchemaService
except ImportError:
    AISchemaService = None  # type: ignore
from src.db.session import async_operation_lock
from src.modules.data.utils.credentials import encrypt_credentials, decrypt_credentials
from src.core.edition import is_ee_enabled
from src.shared.query_limits import (
    DEFAULT_PAGE_LIMIT,
    DEFAULT_FILE_QUERY_LIMIT,
    DEFAULT_LIST_PAGE_LIMIT,
    FILE_UPLOAD_SAMPLE_ROWS,
    UPLOAD_IN_MEMORY_MAX_SAMPLE_ROWS,
    PREVIEW_ROWS,
)

# NoSQL connector imports (optional)
try:
    from src.modules.data.connectors.mongodb_connector import MongoDBConnector
    MONGODB_AVAILABLE = True
except ImportError:
    MONGODB_AVAILABLE = False
    MongoDBConnector = None  # type: ignore[misc, assignment]

try:
    from src.modules.data.connectors.cassandra_connector import CassandraConnector
    CASSANDRA_AVAILABLE = True
except ImportError:
    CASSANDRA_AVAILABLE = False
    CassandraConnector = None  # type: ignore[misc, assignment]

try:
    from src.modules.data.connectors.dynamodb_connector import DynamoDBConnector
    DYNAMODB_AVAILABLE = True
except ImportError:
    DYNAMODB_AVAILABLE = False
    DynamoDBConnector = None  # type: ignore[misc, assignment]

logger = logging.getLogger(__name__)

NOSQL_TYPES = ("mongodb", "cassandra", "dynamodb")
# IoT / time-series sources routed to EnterpriseConnectorsService (EE)
TIMESERIES_TYPES = ("influxdb", "prometheus_source", "opensearch", "elasticsearch")


def _normalize_connection_config(connection_config: Dict[str, Any]) -> Dict[str, Any]:
    """Return a shallow copy of connection_config for safe mutation. Ensures dict shape."""
    if not isinstance(connection_config, dict):
        return {}
    return dict(connection_config)


def _parse_json_field(value: Any) -> Dict[str, Any]:
    """Parse DB JSON field (may be dict or JSON str) to dict."""
    if isinstance(value, dict):
        return dict(value)
    if isinstance(value, str) and value.strip():
        try:
            return json.loads(value)
        except (json.JSONDecodeError, TypeError):
            pass
    return {}


def _parse_google_sheet_url(sheet_url: str, config_gid: Optional[Any] = None) -> tuple:
    """
    Parse Google Sheet URL and return (spreadsheet_id, gid).
    Handles:
      - https://docs.google.com/spreadsheets/d/ID/edit?gid=123#gid=123
      - https://docs.google.com/spreadsheets/d/ID/edit?usp=sharing
    GID is taken from config_gid if provided, else from URL (?gid=, &gid=, or #gid=), else "0".
    """
    url = (sheet_url or "").strip()
    if not url:
        raise ValueError("sheet_url is required")
    # Spreadsheet ID: /d/SPREADSHEET_ID/ (ID is alphanumeric, hyphen, underscore)
    id_match = re.search(r"/d/([a-zA-Z0-9_-]+)", url)
    if not id_match:
        raise ValueError("Invalid Google Sheet URL: could not extract spreadsheet ID (expect /d/ID in path)")
    sheet_id = id_match.group(1)
    # GID: from config, or from URL query/hash (?gid= &gid= #gid=)
    gid = "0"
    if config_gid is not None and str(config_gid).strip() != "":
        gid = str(config_gid).strip()
    else:
        gid_match = re.search(r"[?#&]gid=(\d+)", url, re.IGNORECASE)
        if gid_match:
            gid = gid_match.group(1)
    return (sheet_id, gid)


# In-memory cache for sample_duckdb schema by domain (shared file = same schema for all tenants).
_SAMPLE_DUCKDB_SCHEMA_CACHE: Dict[str, Dict[str, Any]] = {}
_SAMPLE_DUCKDB_SCHEMA_CACHE_TTL_SEC = 600  # 10 minutes


class DataConnectivityService:
    """Service for handling data connectivity and file uploads"""
    
    def __init__(self):
        _max_mb = int(os.getenv("MAX_FILE_SIZE_MB", 300))
        self.max_file_size = _max_mb * 1024 * 1024
        self.supported_formats = ['csv', 'xlsx', 'xls', 'json', 'txt', 'tsv', 'parquet', 'parq', 'snappy']
        # Enhanced file processing options
        self.file_processing_configs = {
            'csv': {'delimiters': [',', ';', '\t', '|', ' '], 'encodings': ['utf-8', 'latin-1', 'cp1252'], 'auto_detect': True},
            'tsv': {'delimiters': ['\t'], 'encodings': ['utf-8', 'latin-1'], 'auto_detect': False},
            'xlsx': {'sheet_selection': 'auto', 'header_row': 0, 'skip_rows': 0},
            'xls': {'sheet_selection': 'auto', 'header_row': 0, 'skip_rows': 0},
            'json': {'flatten_nested': True, 'max_nesting_level': 3},
            'txt': {'line_mode': True},
            'parquet': {'columns': 'auto', 'partitions': 'auto'},
        }
        self.data_sources = {}
        self._initialize_demo_data()
        self.database_connector = DatabaseConnectorService()
        self.ai_schema_service = AISchemaService()
        # Schema cache: {data_source_id: (schema_dict, fetched_at_timestamp)}
        self._schema_cache: Dict[str, tuple] = {}
        self._schema_cache_ttl = 300  # 5 minutes

    def invalidate_data_source_cache(self, data_source_id: str) -> None:
        """Remove a data source from in-memory cache so next read gets fresh config (e.g. after update)."""
        self.data_sources.pop(data_source_id, None)
        self._schema_cache.pop(data_source_id, None)
        logger.debug("Invalidated in-memory cache for data source: %s", data_source_id)

    async def _test_nosql_connection(self, db_type: str, request: Dict[str, Any]) -> Dict[str, Any]:
        """Test NoSQL connection (MongoDB, Cassandra, DynamoDB) using dedicated connectors."""
        import asyncio
        try:
            if db_type == "mongodb" and MONGODB_AVAILABLE and MongoDBConnector:
                conn_str = request.get("connection_string") or request.get("host") or ""
                database = request.get("database") or ""
                if not conn_str or not database:
                    return {"success": False, "error": "MongoDB requires connection_string (or host) and database."}
                connector = MongoDBConnector(conn_str, database)
                ok = await asyncio.to_thread(connector.connect)
                if ok:
                    connector.disconnect()
                return {"success": bool(ok), "message": "MongoDB connection successful" if ok else "MongoDB connection failed", "connection_info": {}}
            if db_type == "cassandra" and CASSANDRA_AVAILABLE and CassandraConnector:
                host_str = request.get("host") or request.get("contact_points") or ""
                contact_points = [x.strip() for x in host_str.split(",") if x.strip()] if isinstance(host_str, str) else host_str
                keyspace = request.get("keyspace") or request.get("database") or ""
                if not contact_points or not keyspace:
                    return {"success": False, "error": "Cassandra requires host (contact points) and keyspace/database."}
                connector = CassandraConnector(
                    contact_points,
                    keyspace,
                    request.get("username"),
                    request.get("password"),
                )
                ok = await asyncio.to_thread(connector.connect)
                if ok:
                    connector.disconnect()
                return {"success": bool(ok), "message": "Cassandra connection successful" if ok else "Cassandra connection failed", "connection_info": {}}
            if db_type == "dynamodb" and DYNAMODB_AVAILABLE and DynamoDBConnector:
                region = request.get("region") or "us-east-1"
                access_key = request.get("access_key_id") or request.get("accessKey")
                secret_key = request.get("secret_access_key") or request.get("secretKey")
                if not access_key or not secret_key:
                    return {"success": False, "error": "DynamoDB requires access_key_id and secret_access_key (or accessKey/secretKey)."}
                connector = DynamoDBConnector(region, access_key, secret_key)
                # DynamoDB has no connect(); test by listing tables (small call)
                try:
                    await asyncio.to_thread(connector.client.list_tables, Limit=1)
                except Exception as e:
                    return {"success": False, "error": str(e)}
                return {"success": True, "message": "DynamoDB connection successful", "connection_info": {}}
            if db_type in NOSQL_TYPES:
                return {
                    "success": False,
                    "error": f"NoSQL type {db_type} is not available (driver not installed or disabled)."
                }
            return {"success": False, "error": f"Unsupported NoSQL type: {db_type}"}
        except Exception as e:
            logger.exception("NoSQL connection test failed")
            return {"success": False, "error": str(e)}

    async def _test_timeseries_connection(self, db_type: str, request: Dict[str, Any]) -> Dict[str, Any]:
        """Test IoT / time-series connection via EnterpriseConnectorsService (plan-gated in product; not AISER_EDITION)."""
        try:
            from ee.modules.data.services.enterprise_connectors_service import (
                EnterpriseConnectorsService, ConnectionConfig, ConnectorType,
            )
            svc = EnterpriseConnectorsService()
            conn_type = ConnectorType(db_type)
            config = ConnectionConfig(
                connector_type=conn_type,
                name=request.get("name") or db_type,
                host=request.get("host") or request.get("prometheus_url") or "",
                port=request.get("port"),
                username=request.get("username"),
                password=request.get("password"),
                token=request.get("influxdb_token") or request.get("token"),
                metadata=request,
            )
            handler = svc.supported_connectors.get(conn_type)
            if not handler:
                return {"success": False, "error": f"No handler for connector type: {db_type}"}
            return await handler(config, test_only=True)
        except Exception as e:
            logger.error(f"❌ IoT connection test failed: {e}")
            return {"success": False, "error": str(e)}

    async def _fetch_nosql_schema(self, db_type: str, config: Dict[str, Any]) -> Dict[str, Any]:
        """Fetch schema from NoSQL store and return normalized { tables, schemas, total_rows }."""
        import asyncio
        try:
            if db_type == "mongodb" and MONGODB_AVAILABLE and MongoDBConnector:
                conn_str = config.get("connection_string") or config.get("host") or ""
                database = config.get("database") or ""
                if not conn_str or not database:
                    return {"success": False, "error": "MongoDB config missing connection_string and database.", "tables": [], "schemas": [], "total_rows": 0}
                connector = MongoDBConnector(conn_str, database)
                await asyncio.to_thread(connector.connect)
                try:
                    raw = await connector.get_schema()
                    connector.disconnect()
                except Exception:
                    connector.disconnect()
                    raise
                tables = []
                for coll_name, coll_info in (raw or {}).items():
                    fields = coll_info.get("fields", [])
                    tables.append({"name": coll_name, "columns": [f.get("name") for f in fields if f.get("name")], "row_count": coll_info.get("total_count", 0)})
                return {"success": True, "tables": tables, "schemas": list(raw.keys()) if isinstance(raw, dict) else [], "total_rows": sum(t.get("row_count", 0) for t in tables)}
            if db_type == "cassandra" and CASSANDRA_AVAILABLE and CassandraConnector:
                host_str = config.get("host") or config.get("contact_points") or ""
                contact_points = [x.strip() for x in host_str.split(",") if x.strip()] if isinstance(host_str, str) else host_str
                keyspace = config.get("keyspace") or config.get("database") or ""
                if not contact_points or not keyspace:
                    return {"success": False, "error": "Cassandra config missing host and keyspace.", "tables": [], "schemas": [], "total_rows": 0}
                connector = CassandraConnector(contact_points, keyspace, config.get("username"), config.get("password"))
                await asyncio.to_thread(connector.connect)
                try:
                    raw = await connector.get_schema()
                    connector.disconnect()
                except Exception:
                    connector.disconnect()
                    raise
                tables = []
                for table_name, table_info in (raw or {}).items():
                    all_cols = table_info.get("all_columns", []) or table_info.get("columns", [])
                    col_names = [c.get("name") if isinstance(c, dict) else str(c) for c in all_cols]
                    tables.append({"name": table_name, "columns": col_names})
                return {"success": True, "tables": tables, "schemas": [keyspace], "total_rows": 0}
            if db_type == "dynamodb" and DYNAMODB_AVAILABLE and DynamoDBConnector:
                region = config.get("region") or "us-east-1"
                access_key = config.get("access_key_id") or config.get("accessKey")
                secret_key = config.get("secret_access_key") or config.get("secretKey")
                if not access_key or not secret_key:
                    return {"success": False, "error": "DynamoDB config missing credentials.", "tables": [], "schemas": [], "total_rows": 0}
                connector = DynamoDBConnector(region, access_key, secret_key)
                table_name = config.get("table_name") or config.get("database")
                if table_name:
                    schema_one = await connector.get_schema(table_name)
                    tables = [{"name": table_name, "columns": list((schema_one or {}).get("attributes", {}).keys())}]
                    return {"success": True, "tables": tables, "schemas": [table_name], "total_rows": (schema_one or {}).get("item_count", 0)}
                # List all tables
                paginator = connector.client.get_paginator("list_tables")
                table_names = []
                for page in paginator.paginate():
                    table_names.extend(page.get("TableNames", []))
                tables = []
                for tn in table_names[:50]:
                    try:
                        sch = await connector.get_schema(tn)
                        tables.append({"name": tn, "columns": list((sch or {}).get("attributes", {}).keys())})
                    except Exception:
                        tables.append({"name": tn, "columns": []})
                return {"success": True, "tables": tables, "schemas": table_names[:50], "total_rows": 0}
            return {"success": False, "error": f"NoSQL schema not implemented for {db_type}.", "tables": [], "schemas": [], "total_rows": 0}
        except Exception as e:
            logger.exception("NoSQL schema fetch failed")
            return {"success": False, "error": str(e), "tables": [], "schemas": [], "total_rows": 0}

    def _make_json_serializable(self, obj: Any) -> Any:
        """Recursively convert date/datetime objects to ISO format strings for JSON serialization"""
        if isinstance(obj, (date, datetime)):
            return obj.isoformat()
        elif isinstance(obj, dict):
            return {key: self._make_json_serializable(value) for key, value in obj.items()}
        elif isinstance(obj, list):
            return [self._make_json_serializable(item) for item in obj]
        elif pd.isna(obj):
            return None
        else:
            return obj

    async def _convert_upload_to_compressed_parquet(
        self,
        source_path: str,
        file_extension: str,
        options: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        """Convert uploaded file to compressed parquet payload for blob storage."""
        options = options or {}

        def _detect_csv_header_and_delimiter(path: str, fallback_delim: str) -> tuple[int, str]:
            candidate_delims = [fallback_delim, ",", ";", "\t", "|"]
            lines: List[str] = []
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                for _, line in zip(range(50), f):
                    lines.append(line.rstrip("\n\r"))

            best_score = -1
            best_header = 0
            best_delim = fallback_delim
            best_fields = 0
            best_consistency = 0
            for delim in candidate_delims:
                for idx, line in enumerate(lines[:30]):
                    fields = line.split(delim) if delim else [line]
                    non_empty = sum(1 for part in fields if str(part).strip())
                    if len(fields) < 2 or non_empty < 2:
                        continue
                    consistency = 0
                    for nxt in lines[idx + 1 : idx + 6]:
                        if len((nxt.split(delim) if delim else [nxt])) == len(fields):
                            consistency += 1
                    score = (len(fields) * 10) + consistency
                    if score > best_score:
                        best_score = score
                        best_header = idx
                        best_delim = delim
                        best_fields = len(fields)
                        best_consistency = consistency

            # Conservative default: first row is the header unless we have strong
            # evidence of preamble text before a tabular header.
            if best_header > 0 and lines:
                first_fields = len(lines[0].split(best_delim)) if best_delim else 1
                if first_fields <= 2 and best_fields >= 4 and best_consistency >= 3:
                    return best_header, best_delim
            return 0, best_delim

        def _read_into_dataframe() -> pd.DataFrame:
            ext = (file_extension or "").lower()
            if ext in ("csv", "tsv"):
                delimiter = options.get("delimiter") or ("\t" if ext == "tsv" else ",")
                header_row = options.get("header_row")
                if header_row is None:
                    header_row, delimiter = _detect_csv_header_and_delimiter(source_path, delimiter)
                return pd.read_csv(
                    source_path,
                    delimiter=delimiter,
                    header=int(header_row),
                    engine="python",
                    on_bad_lines="skip",
                )
            if ext in ("xlsx", "xls"):
                sheet_name = options.get("sheet_name")
                if sheet_name:
                    return pd.read_excel(source_path, sheet_name=sheet_name)
                return pd.read_excel(source_path)
            if ext in ("parquet", "parq", "snappy"):
                return pd.read_parquet(source_path)
            if ext == "json":
                return pd.read_json(source_path)
            if ext == "txt":
                with open(source_path, "r", encoding="utf-8", errors="replace") as f:
                    lines = [line.rstrip("\n\r") for line in f]
                return pd.DataFrame({"text": lines})
            raise ValueError(f"Unsupported upload format for parquet conversion: {ext}")

        df = await asyncio.to_thread(_read_into_dataframe)
        parquet_buffer = io.BytesIO()
        await asyncio.to_thread(
            df.to_parquet,
            parquet_buffer,
            index=False,
            compression="zstd",
        )
        parquet_bytes = parquet_buffer.getvalue()
        return {
            "content": parquet_bytes,
            "storage_format": "parquet",
            "compressed_size_bytes": len(parquet_bytes),
            "row_count": len(df.index),
        }

    async def test_database_connection(self, connection_request: Dict[str, Any]) -> Dict[str, Any]:
        """Test database connection without storing credentials"""
        try:
            logger.info(f"🔌 Testing database connection: {connection_request.get('type')}")
            
            db_type = connection_request.get('type', '').lower()
            
            # Normalize: if uri or connection_string provided, parse into host/port/database/username/password
            # so both "Connection URL" and manual fields work with backend (and SQL Server gets correct params)
            uri_raw = connection_request.get('uri') or connection_request.get('connection_string')
            if uri_raw and isinstance(uri_raw, str) and uri_raw.strip():
                try:
                    parsed = self._parse_database_uri_or_odbc(uri_raw.strip())
                    if parsed:
                        connection_request = {**parsed, **{k: v for k, v in connection_request.items() if v is not None and k not in ('uri', 'connection_string')}}
                        db_type = connection_request.get('type', db_type)
                except Exception as e:
                    logger.warning(f"Failed to parse URI/connection_string: {e}")
            
            # NoSQL: test via dedicated connectors
            if db_type in NOSQL_TYPES:
                return await self._test_nosql_connection(db_type, connection_request)

            # IoT / time-series: test via EnterpriseConnectorsService (EE)
            if db_type in TIMESERIES_TYPES:
                return await self._test_timeseries_connection(db_type, connection_request)

            # Get supported databases from DatabaseConnectorService
            supported_dbs = self.database_connector.get_supported_databases()

            if db_type not in supported_dbs:
                return {
                    'success': False,
                    'error': f'Unsupported database type: {db_type}. Supported: {supported_dbs}'
                }

            # Test connection using DatabaseConnectorService
            try:
                test_result = await self.database_connector.test_connection(connection_request)
                return test_result
                    
            except Exception as e:
                logger.error(f"❌ Database connection test error: {str(e)}")
                return {
                    'success': False,
                    'error': f'Connection test failed: {str(e)}'
                }
                
        except Exception as e:
            logger.error(f"❌ Database connection test failed: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }

    async def store_database_connection(self, connection_request: Dict[str, Any], user_id: str, project_id: str) -> Dict[str, Any]:
        """Store database connection configuration
        
        NOTE: This method validates the connection BEFORE encrypting credentials.
        This ensures the validation works with plain credentials, and also provides
        safety for users who skip the test step and directly click save.
        
        Args:
            connection_request: Database connection configuration
            user_id: User ID (required for data isolation and security)
            project_id: Project ID (required for ownership)
        """
        if not user_id:
            raise ValueError("user_id is required for database connections")

        # EE requires an explicit project_id; CE stores by user_id with no project
        if is_ee_enabled() and not project_id:
            raise ValueError("project_id is required for database connections")
        
        try:
            db_type = str(connection_request.get('type', '')).lower()
            logger.info(f"💾 Storing database connection of type '{db_type}' for user {user_id}")

            # ALWAYS validate the connection first with PLAIN credentials.
            # This prevents persisting phantom data sources when credentials or networking are invalid,
            # and provides safety for users who skip the test step.
            try:
                if db_type in NOSQL_TYPES:
                    test_result = await self._test_nosql_connection(db_type, connection_request)
                elif db_type in TIMESERIES_TYPES:
                    test_result = await self._test_timeseries_connection(db_type, connection_request)
                else:
                    test_result = await self.database_connector.test_connection(connection_request)
                if not test_result.get('success'):
                    error_msg = test_result.get('error') or 'Database connection test failed'
                    logger.warning(f"❌ Connection validation failed; not storing data source: {error_msg}")
                    return {
                        'success': False,
                        'error': error_msg
                    }
            except Exception as validation_error:
                logger.error(f"❌ Unexpected error while validating connection before store: {validation_error}")
                return {
                    'success': False,
                    'error': f'Connection validation failed: {validation_error}'
                }

            # Generate unique ID for the connection
            connection_id = f"db_{connection_request.get('type')}_{int(datetime.now().timestamp())}"
            
            # Prepare in-memory representation upfront to avoid unbound variable
            connection_data = {
                'id': connection_id,
                'type': 'database',
                'db_type': connection_request.get('type'),
                'name': connection_request.get('name', f"{connection_request.get('type')}_connection"),
                'config': connection_request,
                'created_at': datetime.now().isoformat(),
                'status': 'connected',
                'connection_type': 'database'
            }
            
            # Store connection info in database
            try:
                from src.modules.data.models import DataSource
                from src.db.session import async_session

                async with async_session() as db:
                    # Create new data source record
                    # Ensure any sensitive credentials are stored encrypted in DB
                    try:
                        from src.modules.data.utils.credentials import encrypt_credentials
                        safe_config = encrypt_credentials(connection_request)
                        # Log encryption status for debugging
                        if safe_config != connection_request:
                            logger.info(f"✅ Credentials encrypted for {connection_request.get('type')} connection")
                        else:
                            logger.warning(f"⚠️ Credentials not encrypted (ENCRYPTION_KEY may not be set) for {connection_request.get('type')} connection")
                    except Exception as encrypt_error:
                        logger.error(f"❌ Failed to encrypt credentials: {encrypt_error}")
                        safe_config = connection_request
                    
                    # Convert project_id to UUID (nullable)
                    from uuid import UUID
                    project_id_uuid = None
                    if project_id:
                        if isinstance(project_id, str):
                            try:
                                project_id_uuid = UUID(project_id)
                            except ValueError:
                                logger.warning(f"Invalid project_id format: {project_id}, storing without project")
                        else:
                            project_id_uuid = project_id
                    
                    # Resolve user_id UUID so CE list endpoint can find this record by owner
                    user_id_uuid = None
                    try:
                        from uuid import UUID as _UUID
                        user_id_uuid = _UUID(user_id)
                    except (TypeError, ValueError):
                        pass

                    new_source = DataSource(
                        id=connection_id,
                        name=connection_request.get('name') or f"{connection_request.get('type')}_connection",
                        type='database',
                        format=connection_request.get('type'),
                        db_type=connection_request.get('type'),
                        size=0,
                        row_count=0,
                        schema=json.dumps({
                            'type': connection_request.get('type'),
                            'host': connection_request.get('host'),
                            'port': connection_request.get('port'),
                            'database': connection_request.get('database'),
                            'username': connection_request.get('username'),
                            'ssl_mode': connection_request.get('ssl_mode', 'prefer'),
                            'connection_string': connection_request.get('uri'),
                            'encrypt': connection_request.get('encrypt', False)
                        }),
                        connection_config=json.dumps(safe_config),
                        metadata=json.dumps({
                            'connection_type': 'database',
                            'status': 'connected',
                            'created_at': datetime.now().isoformat()
                        }),
                        user_id=user_id_uuid,
                        project_id=project_id_uuid if project_id else None,
                        is_active=True,
                        created_at=datetime.now(),
                        updated_at=datetime.now(),
                        last_accessed=datetime.now()
                    )
                    
                    db.add(new_source)
                    await db.commit()
                    await db.refresh(new_source)
                    
                    logger.info(f"✅ Database connection saved to database: {connection_id}")
                    # Keep in-memory registry aligned
                    connection_data.update({
                        'name': new_source.name,
                        'db_type': new_source.db_type
                    })
            except Exception as db_error:
                logger.error(f"❌ Failed to save to database: {str(db_error)}")
                # Fallback to memory storage (connection_data already prepared)
            self.data_sources[connection_id] = connection_data
            
            # Create connection in DatabaseConnectorService for SQL databases only (NoSQL / time-series use dedicated paths)
            if db_type not in NOSQL_TYPES and db_type not in TIMESERIES_TYPES:
                try:
                    conn_result = await self.database_connector.create_connection(connection_request)
                    if conn_result.get('success'):
                        logger.info(f"✅ Database connection engine created: {connection_id}")
                except Exception as conn_error:
                    logger.warning(f"⚠️ Connection engine creation failed (queries may not work): {str(conn_error)}")
            
            logger.info(f"✅ Database connection stored: {connection_id}")
            
            return {
                'success': True,
                'data_source_id': connection_id,
                'connection_info': {
                    'id': connection_id,
                    'type': 'database',
                    'db_type': connection_request.get('type'),
                    'name': connection_request.get('name', f"{connection_request.get('type')}_connection"),
                    'status': 'connected'
                }
            }
            
        except Exception as e:
            logger.error(f"❌ Failed to store database connection: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }

    async def _read_file_data(self, file_path: str, file_format: str, limit: int = None) -> List[Dict[str, Any]]:
        """Read data from a file based on its format"""
        if limit is None:
            limit = DEFAULT_FILE_QUERY_LIMIT
        try:
            logger.info(f"📁 Reading {file_format} file: {file_path}")
            
            if file_format == 'csv':
                import pandas as pd
                df = pd.read_csv(file_path, nrows=limit)
                return df.to_dict('records')
            elif file_format == 'xlsx':
                import pandas as pd
                df = pd.read_excel(file_path, nrows=limit)
                return df.to_dict('records')
            elif file_format == 'json':
                import json
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                if isinstance(data, list):
                    return data[:limit]
                else:
                    return [data]
            elif file_format == 'txt':
                with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                    lines = [line.rstrip('\n\r') for line in f]
                return [{'text': line} for line in lines[:limit]]
            elif file_format == 'parquet':
                import pandas as pd
                df = pd.read_parquet(file_path)
                return df.head(limit).to_dict('records')
            else:
                logger.warning(f"⚠️ Unsupported file format: {file_format}")
                return []
                
        except Exception as e:
            logger.error(f"❌ Failed to read file {file_path}: {str(e)}")
            return []
    
    async def execute_query_on_source(self, source_id: str, query: str) -> Dict[str, Any]:
        """Execute a custom query on a data source"""
        try:
            logger.info(f"🔍 Executing query on source: {source_id}")
            logger.info(f"🔍 Query: {query}")
            
            # Get the data source
            source = await self.get_data_source_by_id(source_id)
            if not source:
                return {
                    'success': False,
                    'error': f'Data source {source_id} not found'
                }
            
            # For demo data sources, implement basic query parsing
            if source.get('source') == 'demo_data' or source.get('id', '').startswith('demo_'):
                data = source.get('sample_data', [])
                
                # Simple query parsing for demo purposes
                if 'count' in query.lower():
                    # Count query
                    result = len(data)
                    return {
                        'success': True,
                        'data': [{'count': result}],
                        'total_rows': 1,
                        'query_type': 'count'
                    }
                elif 'select' in query.lower() and 'from' in query.lower():
                    # Basic SELECT query simulation
                    # This is a simplified implementation for demo purposes
                    try:
                        # Parse basic SELECT * FROM table
                        if 'select *' in query.lower():
                            return {
                                'success': True,
                                'data': data,
                                'total_rows': len(data),
                                'query_type': 'select_all'
                            }
                        else:
                            # For other queries, return sample data
                            return {
                                'success': True,
                                'data': data[:PREVIEW_ROWS],
                                'total_rows': len(data[:PREVIEW_ROWS]),
                                'query_type': 'select_limited'
                            }
                    except Exception as parse_error:
                        logger.warning(f"Query parsing failed, returning sample data: {parse_error}")
                        return {
                            'success': True,
                            'data': data[:PREVIEW_ROWS],
                            'total_rows': len(data[:PREVIEW_ROWS]),
                            'query_type': 'fallback'
                        }
                else:
                    # Default: return sample data
                    return {
                        'success': True,
                        'data': data,
                        'total_rows': len(data),
                        'query_type': 'default'
                    }
            
            # For file and database sources, use MultiEngineQueryService for execution
            elif source.get('type') in ('file', 'database', 'warehouse', 'sample_duckdb'):
                try:
                    from src.modules.data.services.multi_engine_query_service import MultiEngineQueryService, get_multi_engine_query_service
                    multi = get_multi_engine_query_service()
                    
                    # MultiEngineQueryService handles connection resolution, engine selection, and execution
                    result = await multi.execute_query(query, source)
                    
                    if result.get('success'):
                        return {
                            'success': True,
                            'data': result.get('data', []),
                            'columns': result.get('columns', []),
                            'total_rows': result.get('row_count', len(result.get('data', []))),
                            'engine': result.get('engine'),
                            'query_type': 'custom'
                        }
                    else:
                        return {
                            'success': False,
                            'error': result.get('error', 'Query execution failed')
                        }
                except Exception as query_err:
                    logger.error(f"❌ MultiEngineQueryService execution failed: {str(query_err)}")
                    return {
                        'success': False,
                        'error': f"Query execution failed: {str(query_err)}"
                    }
            
            else:
                return {
                    'success': False,
                    'error': f'Unsupported data source type: {source.get("type")}'
                }
                
        except Exception as e:
            logger.error(f"❌ Failed to execute query on source {source_id}: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    async def get_source_schema(self, source_id: str) -> Dict[str, Any]:
        """Get schema information for a specific data source"""
        try:
            logger.info(f"🔍 Getting schema for source: {source_id}")
            
            # Get the data source
            source = await self.get_data_source_by_id(source_id)
            if not source:
                return {
                    'success': False,
                    'error': f'Data source {source_id} not found'
                }
            
            # Return schema information
            schema = source.get('schema', {})
            # Parse JSON string schemas from database if necessary
            if isinstance(schema, str):
                try:
                    schema = json.loads(schema)
                except json.JSONDecodeError:
                    logger.warning("⚠️ Stored schema is a string but not valid JSON; returning empty schema")
                    schema = {}
            if schema:
                logger.info(f"✅ Returning schema for source {source_id}")
                return {
                    'success': True,
                    'schema': schema,
                    'source_id': source_id,
                    'source_name': source.get('name', 'Unknown'),
                    'source_type': source.get('type', 'Unknown')
                }
            else:
                return {
                    'success': False,
                    'error': f'No schema available for data source {source_id}'
                }
                
        except Exception as e:
            logger.error(f"❌ Failed to get schema for source {source_id}: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    async def get_data_from_source(self, source_id: str, limit: int = None) -> Dict[str, Any]:
        """Get data from a specific data source"""
        if limit is None:
            limit = DEFAULT_PAGE_LIMIT
        try:
            logger.info(f"🔍 Getting data from source: {source_id}")
            
            # Get the data source
            source = await self.get_data_source_by_id(source_id)
            if not source:
                return {
                    'success': False,
                    'error': f'Data source {source_id} not found'
                }
            
            # For demo data sources, return sample data
            if source.get('source') == 'demo_data' or source.get('id', '').startswith('demo_'):
                sample_data = source.get('sample_data', [])
                logger.info(f"✅ Returning {len(sample_data)} sample data rows from demo source")
                return {
                    'success': True,
                    'data': sample_data,
                    'total_rows': len(sample_data),
                    'source': source
                }
            
            # For file sources, read the actual file
            elif source.get('type') == 'file':
                file_path = source.get('file_path')
                if not file_path:
                    return {
                        'success': False,
                        'error': 'No file path available for this data source'
                    }
                
                # Read file data based on format
                file_format = source.get('format', 'csv').lower()
                data = await self._read_file_data(file_path, file_format, limit)
                
                return {
                    'success': True,
                    'data': data,
                    'total_rows': len(data),
                    'source': source
                }
            
            # For database sources, query the database using MultiEngineQueryService
            elif source.get('type') in ('database', 'warehouse', 'sample_duckdb'):
                try:
                    # Resolve table name from schema
                    schema_info = source.get('schema', {})
                    if isinstance(schema_info, str):
                        try:
                            schema_info = json.loads(schema_info)
                        except:
                            schema_info = {}
                    
                    table_name = None
                    schema_name = "public"
                    
                    if schema_info.get('table'):
                        table_name = schema_info.get('table')
                        schema_name = schema_info.get('schema', 'public')
                    elif schema_info.get('tables') and len(schema_info.get('tables')) > 0:
                        # Pick active or first table
                        target = schema_info.get('tables')[0]
                        for t in schema_info.get('tables'):
                            if t.get('active') or t.get('is_active'):
                                target = t
                                break
                        table_name = target.get('name')
                        schema_name = target.get('schema') or schema_info.get('schema') or 'public'
                    
                    if not table_name:
                        return {
                            'success': False,
                            'error': 'Could not identify a table to query in the database schema'
                        }
                    
                    sql = f"SELECT * FROM {schema_name}.{table_name} LIMIT {limit or 100}"
                    
                    from src.modules.data.services.multi_engine_query_service import MultiEngineQueryService, get_multi_engine_query_service
                    multi = get_multi_engine_query_service()
                    result = await multi.execute_query(sql, source)
                    
                    if result.get('success'):
                        return {
                            'success': True,
                            'data': result.get('data', []),
                            'total_rows': result.get('row_count', len(result.get('data', []))),
                            'source': source
                        }
                    else:
                        return {
                            'success': False,
                            'error': result.get('error', 'Database query failed')
                        }
                except Exception as e:
                    logger.error(f"❌ Database data retrieval failed: {str(e)}")
                    return {
                        'success': False,
                        'error': f'Database data retrieval failed: {str(e)}'
                    }
            
            else:
                return {
                    'success': False,
                    'error': f'Unsupported data source type: {source.get("type")}'
                }
                
        except Exception as e:
            logger.error(f"❌ Failed to get data from source {source_id}: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    async def get_data_source_by_id(self, source_id: str) -> Optional[Dict[str, Any]]:
        """Get a specific data source by ID"""
        try:
            # Development aliases: map common demo UI ids to available demo sources
            alias_map = {
                'duckdb_local': 'demo_sales_data',
                'csv_sales': 'demo_sales_data',
                'snowflake_warehouse': 'demo_customers_data',
                'postgresql_prod': 'demo_customers_data',
            }
            source_id = alias_map.get(source_id, source_id)
            # First check in-memory demo sources
            if source_id in self.data_sources:
                logger.info(f"✅ Found data source {source_id} in demo sources")
                return self.data_sources[source_id]
            
            # Then check database
            from src.modules.data.models import DataSource
            from src.db.session import async_session

            # Use async_session() which returns a session that can be used as context manager
            from src.db.session import async_session
            async with async_session() as db:
                # Ensure a per-session operation lock exists when the caller used
                # the async_session factory directly (not via get_async_session).
                # This prevents asyncpg 'another operation is in progress' errors
                # by serializing DB operations within this session.
                try:
                    if not getattr(db, '_op_lock', None):
                        import asyncio as _asyncio

                        db._op_lock = _asyncio.Lock()
                except Exception:
                    db._op_lock = None
                from sqlalchemy import select
                query = select(DataSource).where(DataSource.id == source_id, DataSource.is_active == True)
                result = await db.execute(query)
                source = result.scalar_one_or_none()
                
                if source:
                    source_dict = {
                        'id': source.id,
                        'name': source.name,
                        'type': source.type,
                        'format': source.format,
                        # expose persisted inline/sample data and file path if present
                        'sample_data': getattr(source, 'sample_data', None),
                        'file_path': getattr(source, 'file_path', None),
                        'db_type': source.db_type,
                        'size': source.size,
                        'row_count': source.row_count,
                        'schema': source.schema,
                        'user_id': str(source.user_id) if getattr(source, 'user_id', None) else None,
                        'project_id': str(source.project_id) if getattr(source, 'project_id', None) else None,
                        'tenant_id': str(source.tenant_id) if getattr(source, 'tenant_id', None) else None,
                        'created_at': source.created_at.isoformat() if source.created_at else None,
                        'updated_at': source.updated_at.isoformat() if source.updated_at else None,
                        'is_active': source.is_active,
                        'last_accessed': source.last_accessed.isoformat() if source.last_accessed else None
                    }
                    
                    # CRITICAL: Include connection_config for database/warehouse/API/sample_duckdb/google_sheets and decrypt credentials
                    if (source.type in ('database', 'warehouse', 'api', 'sample_duckdb', 'google_sheets')) and source.connection_config:
                        try:
                            import json
                            config = json.loads(source.connection_config) if isinstance(source.connection_config, str) else source.connection_config
                            
                            # CRITICAL: Decrypt credentials before returning
                            try:
                                from src.modules.data.utils.credentials import decrypt_credentials
                                config = decrypt_credentials(config)
                                logger.debug("Connection config loaded for data source %s", source_id)
                            except Exception as decrypt_error:
                                logger.warning(f"⚠️ Could not decrypt credentials for {source_id} (may not be encrypted): {decrypt_error}")
                            
                            # Add to multiple keys for compatibility with query engine
                            source_dict['connection_config'] = config
                            source_dict['connection_info'] = config
                            source_dict['config'] = config
                            source_dict['metadata'] = config  # Also add as metadata for compatibility
                        except Exception as config_error:
                            logger.error(f"❌ Error parsing connection_config for {source_id}: {config_error}")
                            source_dict['connection_config'] = {}
                            source_dict['connection_info'] = {}
                    
                    # For sample_duckdb: if schema is empty, fetch from shared DuckDB so orchestrator/query have table list
                    if source.type == 'sample_duckdb':
                        stored = source_dict.get('schema')
                        has_tables = isinstance(stored, dict) and bool(stored.get('tables'))
                        if not has_tables:
                            try:
                                schema_result = await self.get_sample_duckdb_schema(source_dict)
                                if schema_result.get('success') and schema_result.get('schema'):
                                    source_dict['schema'] = schema_result['schema']
                                    if schema_result.get('data_source', {}).get('row_count') is not None:
                                        source_dict['row_count'] = schema_result['data_source']['row_count']
                                    logger.info(f"✅ Loaded sample_duckdb schema for {source_id}: {len(schema_result.get('schema', {}).get('tables', []))} tables")
                            except Exception as schema_err:
                                logger.warning(f"⚠️ Could not load sample_duckdb schema for {source_id}: {schema_err}")
                    # For google_sheets: if schema is empty, fetch from sheet URL (CSV export)
                    if source.type == 'google_sheets':
                        stored = source_dict.get('schema')
                        has_tables = isinstance(stored, dict) and bool(stored.get('tables'))
                        if not has_tables:
                            try:
                                schema_result = await self.get_google_sheets_schema(source_dict)
                                if schema_result.get('success') and schema_result.get('schema'):
                                    source_dict['schema'] = schema_result['schema']
                                    if schema_result.get('data_source', {}).get('row_count') is not None:
                                        source_dict['row_count'] = schema_result['data_source']['row_count']
                                    logger.info(f"✅ Loaded google_sheets schema for {source_id}")
                            except Exception as schema_err:
                                logger.warning(f"⚠️ Could not load google_sheets schema for {source_id}: {schema_err}")
                    
                    logger.info(f"✅ Found data source {source_id} in database (type: {source.type}, has_config: {bool(source_dict.get('connection_config'))})")
                    return source_dict
                
                logger.warning(f"⚠️ Data source {source_id} not found in database or demo sources")
                return None
                
        except Exception as e:
            logger.error(f"❌ Failed to get data source {source_id}: {str(e)}")
            # Fallback to in-memory sources
            return self.data_sources.get(source_id)
    
    async def get_data_sources(self, offset: int = 0, limit: int = None) -> List[Dict[str, Any]]:
        """Get all data sources with pagination from database"""
        if limit is None:
            limit = DEFAULT_LIST_PAGE_LIMIT
        try:
            logger.info(f"🔍 Getting data sources (offset: {offset}, limit: {limit})")
            logger.info(f"🔍 Available demo sources: {list(self.data_sources.keys())}")
            
            from src.modules.data.models import DataSource
            from src.db.session import async_session

            async with async_session() as db:
                # Query database for data sources using async session
                from sqlalchemy import select, func
                
                # Count total
                count_query = select(func.count(DataSource.id)).where(DataSource.is_active == True)
                total_result = await db.execute(count_query)
                total = total_result.scalar()
                
                # Get paginated results
                query = select(DataSource).where(DataSource.is_active == True).offset(offset).limit(limit)
                result = await db.execute(query)
                sources = result.scalars().all()
                
                # Convert to dictionary format
                result_list = []
                for source in sources:
                    source_dict = {
                        'id': source.id,
                        'name': source.name,
                        'type': source.type,
                        'format': source.format,
                        'db_type': source.db_type,
                        'size': source.size,
                        'row_count': source.row_count,
                        'schema': source.schema,
                        'created_at': source.created_at.isoformat() if source.created_at else None,
                        'updated_at': source.updated_at.isoformat() if source.updated_at else None,
                        'is_active': source.is_active,
                        'last_accessed': source.last_accessed.isoformat() if source.last_accessed else None
                    }
                    result_list.append(source_dict)
                
                logger.info(f"✅ Retrieved {len(result_list)} data sources from database")
                
                # Always include demo data sources for testing
                demo_sources = list(self.data_sources.values())
                logger.info(f"🔍 Adding {len(demo_sources)} demo sources")
                all_sources = result_list + demo_sources
                logger.info(f"🔍 Total sources after combining: {len(all_sources)}")
                
                # Apply pagination to combined sources
                result = all_sources[offset:offset + limit]
                logger.info(f"🔍 Returning {len(result)} sources after pagination")
                return result
                
        except Exception as e:
            logger.error(f"❌ Failed to get data sources from database: {str(e)}")
            import traceback
            logger.error(f"❌ Traceback: {traceback.format_exc()}")
            # Fallback to in-memory sources
            sources = list(self.data_sources.values())
            logger.info(f"✅ Using {len(sources)} in-memory demo data sources")
            return sources[offset:offset + limit]

    async def _save_data_source_to_db(self, data_source: Dict[str, Any]) -> bool:
        """Save data source to database"""
        try:
            from src.modules.data.models import DataSource
            from src.db.session import async_session

            async with async_session() as db:
                # Check if data source already exists
                from sqlalchemy import select
                existing_query = select(DataSource).where(DataSource.id == data_source['id'])
                existing_result = await db.execute(existing_query)
                existing = existing_result.scalar_one_or_none()
                
                if existing:
                    # Update existing
                    existing.name = data_source['name']
                    existing.type = data_source['type']
                    existing.format = data_source.get('format')
                    existing.size = data_source.get('size')
                    existing.row_count = data_source.get('row_count')
                    existing.schema = data_source.get('schema')
                    existing.file_path = data_source.get('file_path')  # Now stores object_key
                    existing.original_filename = data_source.get('original_filename')
                    existing.sample_data = data_source.get('sample_data')  # Save sample_data as JSON
                    from datetime import timezone
                    existing.updated_at = datetime.now(timezone.utc)
                    logger.info(f"✅ Updated data source {data_source['id']} in database.")
                else:
                    # Create new
                    # Use project_id from data_source dict (passed from options)
                    project_id = data_source.get('project_id')
                    
                    if not project_id and is_ee_enabled():
                        raise ValueError(f"project_id is required when creating data source {data_source.get('id')}. Ensure project_id is passed in options.")
                    
                    # Convert project_id to UUID if it's a string
                    from uuid import UUID
                    if isinstance(project_id, str):
                        try:
                            project_id = UUID(project_id)
                        except ValueError:
                            # If conversion fails, try converting integer ID to UUID
                            # This handles legacy numeric IDs that may be sent
                            raise ValueError(f"Invalid project_id format: {project_id}. Must be a valid UUID.")
                    
                    # Normalize user_id to UUID (nullable)
                    user_id_str = data_source.get('user_id')
                    user_id_uuid = None
                    if user_id_str:
                        try:
                            user_id_uuid = UUID(user_id_str)
                        except (ValueError, AttributeError):
                            logger.warning(f"Could not convert user_id '{user_id_str}' to UUID, skipping")

                    new_data_source = DataSource(
                        id=data_source['id'],
                        name=data_source['name'],
                        type=data_source['type'],
                        format=data_source.get('format'),
                        size=data_source.get('size'),
                        row_count=data_source.get('row_count'),
                        schema=data_source.get('schema'),
                        file_path=data_source.get('file_path'),  # Now stores object_key
                        original_filename=data_source.get('original_filename'),
                        sample_data=data_source.get('sample_data'),  # Save sample_data as JSON
                        project_id=project_id,  # Required - validated above
                        user_id=user_id_uuid  # Track who uploaded this data source
                    )
                    db.add(new_data_source)
                    logger.info(f"✅ Saved new data source {data_source['id']} to database (project_id={project_id}).")
                
                await db.commit()
                return True
                
        except Exception as e:
            logger.error(f"❌ Failed to save data source to database: {str(e)}", exc_info=True)
            import traceback
            logger.error(f"❌ Traceback: {traceback.format_exc()}")
            return False

    async def process_uploaded_file(
        self, 
        file_path: str,  # Temp file path for processing
        original_filename: str, 
        options: Optional[Dict[str, Any]] = None,
        object_key: Optional[str] = None  # Object key from datasource storage
    ) -> Dict[str, Any]:
        """Process uploaded file and extract data"""
        try:
            logger.info(f"📁 Processing uploaded file: {original_filename}")
            
            if options is None:
                options = {}
            
            file_extension = self._get_file_extension(original_filename)
            
            if file_extension not in self.supported_formats:
                raise ValueError(f"Unsupported file format: {file_extension}")
            
            # Process file based on extension
            if file_extension == 'csv':
                data, schema = await self._process_csv_file(
                    file_path,
                    options.get('delimiter', ','),
                    options.get('encoding', 'utf-8'),
                    options.get('header_row'),
                )
            elif file_extension == 'tsv':
                data, schema = await self._process_csv_file(
                    file_path,
                    '\t',
                    'utf-8',
                    options.get('header_row'),
                )
            elif file_extension == 'parquet':
                data, schema = await self._process_parquet_file(file_path)
            elif file_extension == 'json':
                data, schema = await self._process_json_file(file_path)
            elif file_extension == 'txt':
                data, schema = await self._process_text_file(file_path)
            elif file_extension in ['xlsx', 'xls']:
                data, schema = await self._process_excel_file(file_path, options.get('sheet_name'))
            else:
                raise ValueError(f"Unsupported file format: {file_extension}")
            
            # Enhance schema with AI insights (skip for preview-only or if disabled for performance)
            preview_only = options.get('preview_only', False)
            enhanced_schema = schema
            # skip_ai_enhancement = preview_only or options.get('skip_ai_enhancement', False)
            
            # if skip_ai_enhancement:
            #     # Use basic schema without AI enhancement for faster processing
            #     enhanced_schema = schema
            #     logger.info("⏩ Skipping AI schema enhancement for faster processing")
            # else:
            #     # Only enhance schema for a sample of data to improve performance
            #     sample_size = min(100, len(data))  # Use first 100 rows for schema enhancement
            #     sample_data = data[:sample_size] if data else []
            #     enhanced_schema = await self.ai_schema_service.generate_enhanced_schema(
            #         sample_data, schema, original_filename, "file"
            #     )
            #     logger.info(f"✅ AI schema enhancement completed using {sample_size} sample rows")
            
            # Generate data source metadata
            user_id = options.get('user_id') if options else None
            project_id = options.get('project_id') if options else None
            name = options.get('name') if options else original_filename
            
            # Get file size from temp file
            file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
            
            # Use schema row_count (full file) not len(data) (sample size)
            total_row_count = enhanced_schema.get('row_count')
            if total_row_count is None:
                total_row_count = len(data)
            data_source = {
                'id': options.get('source_id') or f"file_{int(datetime.now().timestamp())}",
                'name': name or original_filename,
                'type': 'file',
                'format': file_extension,
                'size': file_size,
                'uploaded_at': datetime.now().isoformat(),
                'schema': enhanced_schema,
                'row_count': total_row_count,
                'file_path': object_key,  # NEW: Store object_key, not file path
                'original_filename': original_filename,
                'preview': data[:PREVIEW_ROWS] if len(data) > PREVIEW_ROWS else data,
                # Add fields expected by frontend
                'uuid_filename': f"file_{int(datetime.now().timestamp())}_{original_filename}",
                'content_type': f"application/{file_extension}",
                'storage_type': options.get('storage_type', 'postgresql'),
                'user_id': user_id,  # Pass user_id from options for audit
                'project_id': project_id  # Pass project_id from options (REQUIRED)
            }

            storage_meta = options.get("storage_meta") or {}
            if isinstance(storage_meta, dict) and storage_meta:
                data_source["schema"] = data_source.get("schema") or {}
                if isinstance(data_source["schema"], dict):
                    storage_info = {
                        "backend": storage_meta.get("backend", options.get("storage_type", "postgresql")),
                        "format": storage_meta.get("storage_format", "parquet"),
                        "uploaded_size_bytes": storage_meta.get("uploaded_size_bytes", file_size),
                        "stored_size_bytes": storage_meta.get("compressed_size_bytes", file_size),
                    }
                    if storage_meta.get("compression"):
                        storage_info["compression"] = storage_meta.get("compression")
                    data_source["schema"]["storage"] = storage_info
            
            # Conditional in-memory storage based on upload_with_prompt flag
            upload_with_prompt = options.get('upload_with_prompt', False)
            max_sample_rows = UPLOAD_IN_MEMORY_MAX_SAMPLE_ROWS

            if upload_with_prompt:
                # Store in memory for immediate query processing
                if len(data) > max_sample_rows:
                    data_source['sample_data'] = data[:max_sample_rows]
                    data_source['data'] = data[:max_sample_rows]
                else:
                    data_source['data'] = data
                    data_source['sample_data'] = data

                # Store in memory cache
                self.data_sources[data_source['id']] = data_source
                logger.info(f"💾 Stored in memory for immediate processing")
            else:
                # Don't store in memory - only save to DB
                data_source['sample_data'] = data[:max_sample_rows] if len(data) > max_sample_rows else data
                logger.info(f"⏩ Skipping in-memory storage (file uploaded without prompt)")
            
            # Always save to database
            if not preview_only:
                save_success = await self._save_data_source_to_db(data_source)
                if not save_success:
                    logger.error(f"❌ Failed to save data source {data_source.get('id')} to database")
                    raise Exception(f"Failed to save data source to database. Check logs for details.")
            
            logger.info(f"✅ File processed successfully: {len(data)} rows, {len(enhanced_schema['columns'])} columns")
            
            return {
                'success': True,
                'data_source': data_source,
                'data': data if options.get('include_data') else None
            }
            
        except Exception as error:
            logger.error(f"❌ File processing failed: {str(error)}")
            
            return {
                'success': False,
                'error': str(error)
            }

    async def upload_file(
        self,
        file_content: bytes,
        filename: str,
        options: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """Upload and process a file from content"""
        tmp_file_path = None
        try:
            logger.info(f"📁 File upload request: {filename}")
            
            if options is None:
                options = {}
            
            # Validate file size
            if len(file_content) > self.max_file_size:
                raise ValueError(f"File too large. Maximum size: {self.max_file_size / (1024*1024):.1f}MB")
            
            # Validate file format
            file_extension = self._get_file_extension(filename)
            if file_extension not in self.supported_formats:
                raise ValueError(f"Unsupported file format: {file_extension}")
            
            # Get project_id from options
            project_id = options.get('project_id')
            if not project_id and is_ee_enabled():
                raise ValueError("project_id is required for file upload")
            
            # Create temp file for processing (will be deleted after processing)
            with tempfile.NamedTemporaryFile(delete=False, suffix=f".{file_extension}") as tmp_file:
                tmp_file.write(file_content)
                tmp_file_path = tmp_file.name

            # Excel files (.xlsx/.xls) must be stored as-is so all sheets remain accessible at
            # query time via _load_excel_all_sheets_into_duckdb. Converting to parquet would
            # discard every sheet after the first.
            is_excel = file_extension in ("xlsx", "xls")

            if is_excel:
                stored_content = file_content
                stored_filename = filename
                stored_format = file_extension
                stored_content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                stored_size = len(file_content)
            else:
                parquet_payload = await self._convert_upload_to_compressed_parquet(
                    source_path=tmp_file_path,
                    file_extension=file_extension,
                    options=options,
                )
                stored_content = parquet_payload["content"]
                stored_filename = f"{Path(filename).stem}.parquet"
                stored_format = "parquet"
                stored_content_type = "application/x-parquet"
                stored_size = parquet_payload["compressed_size_bytes"]

            # Pre-generate source_id so it can be used in both the storage path and the data source record
            import uuid as _uuid
            source_id = str(_uuid.uuid4())

            storage_service = UploadDatasourceStorageService()
            object_key = await storage_service.store_file(
                file_content=stored_content,
                project_id=project_id,
                original_filename=stored_filename,
                content_type=stored_content_type,
                source_id=source_id,
                organization_id=options.get("organization_id"),
                user_id=options.get("user_id"),
            )
            logger.info(
                "💾 Stored %s in %s: %s (uploaded=%s bytes, stored=%s bytes)",
                stored_format,
                storage_service.storage_type,
                object_key,
                len(file_content),
                stored_size,
            )

            options_with_storage = {
                **options,
                "source_id": source_id,
                "storage_type": storage_service.storage_type,
                "storage_meta": {
                    "backend": storage_service.storage_type,
                    "storage_format": stored_format,
                    "uploaded_size_bytes": len(file_content),
                    "compressed_size_bytes": stored_size,
                    **({"compression": "zstd"} if not is_excel else {}),
                },
            }

            # Process the uploaded file (pass object_key)
            result = await self.process_uploaded_file(tmp_file_path, filename, options_with_storage, object_key)
            
            if result['success']:
                logger.info(f"✅ File upload completed successfully: {filename}")
                return result
            else:
                return result
                
        except Exception as error:
            logger.error(f"❌ File upload failed: {str(error)}")
            return {
                'success': False,
                'error': str(error)
            }
        finally:
            # Clean up temp file
            if tmp_file_path and os.path.exists(tmp_file_path):
                try:
                    os.unlink(tmp_file_path)
                except Exception as e:
                    logger.warning(f"⚠️ Failed to clean up temp file: {str(e)}")

    async def upload_file_from_path(
        self,
        tmp_file_path: str,
        filename: str,
        file_size: int,
        options: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        """Upload and process a file that has already been streamed to a temp path.

        Avoids holding the entire file in memory — the caller streams directly to disk
        and passes the path here. The temp file is NOT deleted by this method; the
        caller owns cleanup.
        """
        if options is None:
            options = {}

        logger.info("📁 File upload request (from path): %s (%d bytes)", filename, file_size)

        if file_size > self.max_file_size:
            raise ValueError(f"File too large. Maximum size: {self.max_file_size / (1024 * 1024):.0f}MB")

        file_extension = self._get_file_extension(filename)
        if file_extension not in self.supported_formats:
            raise ValueError(f"Unsupported file format: {file_extension}")

        project_id = options.get("project_id")
        if not project_id and is_ee_enabled():
            raise ValueError("project_id is required for file upload")

        is_excel = file_extension in ("xlsx", "xls")

        if is_excel:
            with open(tmp_file_path, "rb") as source_file:
                stored_content = source_file.read()
            stored_filename = filename
            stored_format = file_extension
            stored_content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            stored_size = file_size
            storage_compression = None
        else:
            parquet_payload = await self._convert_upload_to_compressed_parquet(
                source_path=tmp_file_path,
                file_extension=file_extension,
                options=options,
            )
            stored_content = parquet_payload["content"]
            stored_filename = f"{Path(filename).stem}.parquet"
            stored_format = "parquet"
            stored_content_type = "application/x-parquet"
            stored_size = parquet_payload["compressed_size_bytes"]
            storage_compression = "zstd"

        import uuid as _uuid
        source_id = str(_uuid.uuid4())

        storage_service = UploadDatasourceStorageService()
        object_key = await storage_service.store_file(
            file_content=stored_content,
            project_id=project_id,
            original_filename=stored_filename,
            content_type=stored_content_type,
            source_id=source_id,
            organization_id=options.get("organization_id"),
            user_id=options.get("user_id"),
        )
        logger.info(
            "💾 Stored %s in %s: %s (original=%d bytes, stored=%d bytes)",
            stored_format,
            storage_service.storage_type,
            object_key,
            file_size,
            stored_size,
        )

        options_with_storage = {
            **options,
            "source_id": source_id,
            "storage_type": storage_service.storage_type,
            "storage_meta": {
                "backend": storage_service.storage_type,
                "storage_format": stored_format,
                "uploaded_size_bytes": file_size,
                "compressed_size_bytes": stored_size,
                **({"compression": storage_compression} if storage_compression else {}),
            },
        }

        return await self.process_uploaded_file(tmp_file_path, filename, options_with_storage, object_key)

    async def _process_csv_file(
        self,
        file_path: str,
        delimiter: str = ',',
        encoding: str = 'utf-8',
        header_row: Optional[int] = None,
    ) -> tuple:
        """Process CSV/TSV files using DuckDB for fast, direct processing"""
        try:
            import duckdb

            def _detect_header_row(path: str, fallback_delim: str) -> tuple[int, str]:
                candidate_delims = [fallback_delim, ",", ";", "\t", "|"]
                lines: List[str] = []
                with open(path, 'r', encoding=encoding, errors='replace') as f:
                    for _, line in zip(range(50), f):
                        lines.append(line.rstrip('\n\r'))
                best_score = -1
                best_row = 0
                best_delim = fallback_delim
                best_fields = 0
                best_consistency = 0
                for delim in candidate_delims:
                    for idx, line in enumerate(lines[:30]):
                        parts = line.split(delim) if delim else [line]
                        non_empty = sum(1 for p in parts if str(p).strip())
                        if len(parts) < 2 or non_empty < 2:
                            continue
                        consistency = 0
                        for nxt in lines[idx + 1: idx + 6]:
                            if len((nxt.split(delim) if delim else [nxt])) == len(parts):
                                consistency += 1
                        score = (len(parts) * 10) + consistency
                        if score > best_score:
                            best_score = score
                            best_row = idx
                            best_delim = delim
                            best_fields = len(parts)
                            best_consistency = consistency

                if best_row > 0 and lines:
                    first_fields = len(lines[0].split(best_delim)) if best_delim else 1
                    if first_fields <= 2 and best_fields >= 4 and best_consistency >= 3:
                        return best_row, best_delim
                return 0, best_delim
            
            # Try to auto-detect delimiter if not specified
            if delimiter == ',':
                detected_delimiter = self._auto_detect_delimiter(file_path)
                if detected_delimiter:
                    delimiter = detected_delimiter
                    logger.info(f"🔍 Auto-detected delimiter: '{delimiter}'")
            if header_row is None:
                header_row, delimiter = _detect_header_row(file_path, delimiter)
                logger.info(f"🔍 Auto-detected CSV header row: {header_row}, delimiter: '{delimiter}'")
            else:
                header_row = max(0, int(header_row))
                logger.info(f"🔍 Using user-provided CSV header row: {header_row}, delimiter: '{delimiter}'")
            
            # Use DuckDB for direct CSV reading (10-100x faster than Pandas)
            conn = duckdb.connect()
            
            # Read CSV directly into DuckDB (auto-detects types, handles encoding)
            try:
                # DuckDB's read_csv_auto uses 'delim' or 'sep', NOT 'delimiter'
                # Escape single quotes in file path for SQL safety
                safe_file_path = file_path.replace("'", "''")
                conn.execute(f"""
                    CREATE TABLE data AS 
                    SELECT * FROM read_csv_auto('{safe_file_path}', 
                        delim='{delimiter}',
                        skip={int(header_row)},
                        header=true,
                        auto_detect=true
                    )
                # """)
                
                # Get schema from DuckDB (faster and more accurate)
                schema_result = conn.execute("DESCRIBE data").fetchall()
                schema = {
                    'columns': [{'name': col[0], 'type': col[1]} for col in schema_result],
                    'row_count': conn.execute("SELECT COUNT(*) FROM data").fetchone()[0]
                }
                
                # Get sample data for preview/schema (unified limit)
                sample_result = conn.execute(f"SELECT * FROM data LIMIT {FILE_UPLOAD_SAMPLE_ROWS}").fetchall()
                columns = [col[0] for col in schema_result]
                
                # Convert to list of dictionaries
                data = [dict(zip(columns, row)) for row in sample_result]
                
                # Convert date/datetime objects to JSON-serializable strings
                data = self._make_json_serializable(data)
                
                # Get full row count
                total_rows = conn.execute("SELECT COUNT(*) FROM data").fetchone()[0]
                schema['row_count'] = total_rows
                
                conn.close()
                
                logger.info(f"🦆 Processed CSV with DuckDB: {total_rows} rows, {len(columns)} columns")
                
                return data, schema
                
            except Exception as duckdb_error:
                logger.warning(f"⚠️ DuckDB CSV read failed, falling back to Pandas: {duckdb_error}")
                conn.close()
                # Fallback to Pandas if DuckDB fails
                df = pd.read_csv(
                    file_path,
                    delimiter=delimiter,
                    encoding=encoding,
                    header=int(header_row),
                    engine='python',
                    on_bad_lines='skip',
                )
                data = df.to_dict('records')
                for row in data:
                    for key, value in row.items():
                        if pd.isna(value):
                            row[key] = None
                
                # Convert date/datetime objects to JSON-serializable strings
                data = self._make_json_serializable(data)
                
                schema = self._infer_schema_from_dataframe(df)
                return data, schema
            
        except Exception as error:
            raise Exception(f"CSV processing failed: {str(error)}")

    async def _process_parquet_file(self, file_path: str) -> tuple:
        """Process Parquet files using DuckDB for native, fast processing"""
        try:
            import duckdb
            
            # Use DuckDB for direct Parquet reading (native format, fastest)
            conn = duckdb.connect()
            
            try:
                # Read Parquet directly into DuckDB (native format, no conversion needed)
                conn.execute(f"CREATE TABLE data AS SELECT * FROM read_parquet('{file_path}')")
                
                # Get schema from DuckDB
                schema_result = conn.execute("DESCRIBE data").fetchall()
                schema = {
                    'columns': [{'name': col[0], 'type': col[1]} for col in schema_result],
                    'row_count': conn.execute("SELECT COUNT(*) FROM data").fetchone()[0]
                }
                
                # Get sample data for preview/schema (unified limit)
                sample_result = conn.execute(f"SELECT * FROM data LIMIT {FILE_UPLOAD_SAMPLE_ROWS}").fetchall()
                columns = [col[0] for col in schema_result]

                # Convert to list of dictionaries
                data = [dict(zip(columns, row)) for row in sample_result]

                # Convert date/datetime objects to JSON-serializable strings
                data = self._make_json_serializable(data)

                # Get full row count
                total_rows = conn.execute("SELECT COUNT(*) FROM data").fetchone()[0]
                schema['row_count'] = total_rows

                # Try to get Parquet metadata
                try:
                    import pyarrow.parquet as pq
                    parquet_file = pq.ParquetFile(file_path)
                    metadata = parquet_file.metadata
                    schema['parquet_metadata'] = {
                        'row_groups': getattr(metadata, 'num_row_groups', 0),
                        'total_rows': total_rows,
                        'created_by': getattr(metadata, 'created_by', 'unknown'),
                        'schema_version': getattr(metadata, 'schema_version', 'unknown')
                    }
                except Exception:
                    schema['parquet_metadata'] = {
                        'row_groups': 1,
                        'total_rows': total_rows,
                        'created_by': 'unknown',
                        'schema_version': 'unknown'
                    }
                
                conn.close()
                
                logger.info(f"🦆 Processed Parquet with DuckDB: {total_rows} rows, {len(columns)} columns")
                
                return data, schema
                
            except Exception as duckdb_error:
                logger.warning(f"⚠️ DuckDB Parquet read failed, falling back to PyArrow: {duckdb_error}")
                conn.close()
                # Fallback to PyArrow/Pandas if DuckDB fails
                import pyarrow.parquet as pq
                df = pq.read_table(file_path).to_pandas()
                data = df.to_dict('records')
                for row in data:
                    for key, value in row.items():
                        if pd.isna(value):
                            row[key] = None
                schema = self._infer_schema_from_dataframe(df)
                return data, schema
            
        except Exception as error:
            raise Exception(f"Parquet processing failed: {str(error)}")

    async def _process_excel_file(self, file_path: str, sheet_name: Optional[str] = None) -> tuple:
        """
        Process Excel files with multi-sheet support using DuckDB.
        Creates virtual tables for each sheet, enabling SQL queries across sheets.
        """
        try:
            import duckdb
            import openpyxl  # For reading Excel sheet names
            import re
            
            # Use DuckDB for Excel processing (supports multi-sheet)
            conn = duckdb.connect()
            
            try:
                # Get all sheet names from Excel file
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                    sheet_names = wb.sheetnames
                    wb.close()
                except Exception as e:
                    logger.warning(f"Could not read Excel sheet names with openpyxl: {e}, using default")
                    # Fallback: try to read with pandas to get sheet names
                    try:
                        xl_file = pd.ExcelFile(file_path)
                        sheet_names = xl_file.sheet_names
                    except:
                        sheet_names = ['Sheet1']  # Default sheet name
                
                logger.info(f"📊 Found {len(sheet_names)} sheet(s) in Excel file: {sheet_names}")
                
                # Process each sheet and create a virtual table
                all_schemas = {}
                primary_data = None
                primary_schema = None
                
                for idx, sheet in enumerate(sheet_names):
                    try:
                        # Read sheet into pandas DataFrame
                        df = pd.read_excel(file_path, sheet_name=sheet, engine='openpyxl')
                        
                        # Clean data (convert NaN to None)
                        df = df.replace({pd.NA: None, pd.NaT: None})
                        df = df.where(pd.notnull(df), None)
                        
                        # Create sanitized table name (DuckDB-safe)
                        table_name = f"sheet_{idx}_{sheet.replace(' ', '_').replace('-', '_').replace('.', '_')[:50]}"
                        # Remove special characters that might break SQL
                        table_name = re.sub(r'[^a-zA-Z0-9_]', '_', table_name)
                        
                        # Register DataFrame in DuckDB
                        conn.register(f"_temp_{table_name}", df)
                        
                        # Create persistent table in DuckDB
                        conn.execute(f"CREATE TABLE {table_name} AS SELECT * FROM _temp_{table_name}")
                        
                        # Get schema for this sheet
                        schema_result = conn.execute(f"DESCRIBE {table_name}").fetchall()
                        sheet_schema = {
                            'columns': [{'name': col[0], 'type': col[1]} for col in schema_result],
                            'row_count': conn.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
                        }
                        
                        all_schemas[sheet] = {
                            'table_name': table_name,
                            'schema': sheet_schema,
                            'row_count': sheet_schema['row_count']
                        }
                        
                        # Use first sheet as primary data (for backward compatibility)
                        if idx == 0 or (sheet_name and sheet == sheet_name):
                            primary_data = df.to_dict('records')
                            # Convert date/datetime objects to JSON-serializable strings
                            primary_data = self._make_json_serializable(primary_data)
                            primary_schema = self._infer_schema_from_dataframe(df)
                            primary_schema['row_count'] = sheet_schema['row_count']
                            primary_schema['table_name'] = table_name
                            primary_schema['all_sheets'] = all_schemas
                        
                        logger.info(f"✅ Created DuckDB table '{table_name}' for sheet '{sheet}' ({sheet_schema['row_count']} rows)")
                        
                    except Exception as sheet_error:
                        logger.warning(f"⚠️ Failed to process sheet '{sheet}': {sheet_error}")
                        continue
                
                if not primary_data:
                    raise Exception("No sheets could be processed from Excel file")

                total_row_count = sum(
                    int(info.get('row_count') or 0)
                    for info in all_schemas.values()
                    if isinstance(info, dict)
                )

                # Build multi-table schema: each sheet becomes a named table
                primary_schema['tables'] = [
                    {
                        'name': sheet,
                        'columns': all_schemas[sheet]['schema']['columns'],
                        'row_count': all_schemas[sheet]['row_count'],
                    }
                    for sheet in sheet_names
                    if sheet in all_schemas
                ]
                primary_schema['all_sheets'] = all_schemas

                # Store DuckDB connection info in schema for later use
                primary_schema['duckdb_tables'] = {sheet: info['table_name'] for sheet, info in all_schemas.items()}
                primary_schema['duckdb_connection'] = 'in_memory'  # Mark that tables are in DuckDB
                primary_schema['primary_table'] = primary_schema.get('table_name')
                primary_schema['primary_row_count'] = primary_schema.get('row_count')
                primary_schema['row_count'] = total_row_count
                primary_schema['total_rows'] = total_row_count
                
                conn.close()
                
                logger.info(f"🦆 Processed Excel with DuckDB: {len(sheet_names)} sheets, total: {total_row_count} rows")
                
                return primary_data, primary_schema
                
            except Exception as duckdb_error:
                logger.warning(f"⚠️ DuckDB Excel processing failed, falling back to Pandas: {duckdb_error}")
                conn.close()
                # Fallback to original Pandas approach
                if sheet_name:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                else:
                    df = pd.read_excel(file_path)
                
                data = df.to_dict('records')
                for row in data:
                    for key, value in row.items():
                        if pd.isna(value):
                            row[key] = None
                
                schema = self._infer_schema_from_dataframe(df)
                return data, schema
            
        except Exception as error:
            raise Exception(f"Excel processing failed: {str(error)}")

    async def _process_json_file(self, file_path: str) -> tuple:
        """Process JSON files using DuckDB for fast, direct processing"""
        try:
            import duckdb
            
            # Use DuckDB for direct JSON reading
            conn = duckdb.connect()
            
            try:
                # Read JSON directly into DuckDB (handles nested structures)
                conn.execute(f"CREATE TABLE data AS SELECT * FROM read_json_auto('{file_path}')")
                
                # Get schema from DuckDB
                schema_result = conn.execute("DESCRIBE data").fetchall()
                schema = {
                    'columns': [{'name': col[0], 'type': col[1]} for col in schema_result],
                    'row_count': conn.execute("SELECT COUNT(*) FROM data").fetchone()[0]
                }
                
                # Get sample data for preview/schema (unified limit)
                sample_result = conn.execute(f"SELECT * FROM data LIMIT {FILE_UPLOAD_SAMPLE_ROWS}").fetchall()
                columns = [col[0] for col in schema_result]

                # Convert to list of dictionaries
                data = [dict(zip(columns, row)) for row in sample_result]

                # Convert date/datetime objects to JSON-serializable strings
                data = self._make_json_serializable(data)

                # Get full row count
                total_rows = conn.execute("SELECT COUNT(*) FROM data").fetchone()[0]
                schema['row_count'] = total_rows

                conn.close()

                logger.info(f"🦆 Processed JSON with DuckDB: {total_rows} rows, {len(columns)} columns")
                
                return data, schema
                
            except Exception as duckdb_error:
                logger.warning(f"⚠️ DuckDB JSON read failed, falling back to Pandas: {duckdb_error}")
                conn.close()
                # Fallback to Pandas if DuckDB fails
                with open(file_path, 'r', encoding='utf-8') as f:
                    import json
                    json_data = json.load(f)
                
                if isinstance(json_data, list):
                    data = json_data[:FILE_UPLOAD_SAMPLE_ROWS]
                    # Convert to DataFrame for schema inference
                    df = pd.DataFrame(data)
                    schema = self._infer_schema_from_dataframe(df)
                else:
                    # Single object - wrap in list
                    data = [json_data]
                    df = pd.DataFrame(data)
                    schema = self._infer_schema_from_dataframe(df)
                
                return data, schema
            
        except Exception as error:
            raise Exception(f"JSON processing failed: {str(error)}")
    
    async def _process_json_file_old(self, file_path: str) -> tuple:
        """Process JSON files"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                json_data = json.load(file)
            
            # Handle different JSON structures
            if isinstance(json_data, list):
                data = json_data
            elif isinstance(json_data, dict):
                if 'data' in json_data and isinstance(json_data['data'], list):
                    data = json_data['data']
                else:
                    data = [json_data]
            else:
                raise ValueError('Invalid JSON structure - expected array or object with data array')
            
            # Convert to DataFrame for schema inference
            df = pd.DataFrame(data)
            
            # Infer schema
            schema = self._infer_schema_from_dataframe(df)
            
            return data, schema
            
        except Exception as error:
            raise Exception(f"JSON processing failed: {str(error)}")

    async def _process_text_file(self, file_path: str) -> tuple:
        """Process text files as one-column datasets."""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='replace') as file:
                lines = [line.rstrip('\n\r') for line in file]
            schema = {
                'columns': [{'name': 'text', 'type': 'VARCHAR'}],
                'row_count': len(lines),
            }
            data = [{'text': line} for line in lines[:FILE_UPLOAD_SAMPLE_ROWS]]
            return data, schema
        except Exception as error:
            raise Exception(f"Text processing failed: {str(error)}")

    def _infer_schema_from_dataframe(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Infer schema from pandas DataFrame"""
        columns = []
        types = {}
        statistics = {}
        
        for column in df.columns:
            series = df[column]
            
            # Determine data type
            if pd.api.types.is_numeric_dtype(series):
                if pd.api.types.is_integer_dtype(series):
                    data_type = 'integer'
                else:
                    data_type = 'number'
                
                # Calculate statistics for numeric columns
                statistics[column] = {
                    'min': float(series.min()) if not pd.isna(series.min()) else None,
                    'max': float(series.max()) if not pd.isna(series.max()) else None,
                    'mean': float(series.mean()) if not pd.isna(series.mean()) else None,
                    'null_count': int(series.isnull().sum())
                }
            elif pd.api.types.is_datetime64_any_dtype(series):
                data_type = 'date'
                statistics[column] = {
                    'null_count': int(series.isnull().sum())
                }
            elif pd.api.types.is_bool_dtype(series):
                data_type = 'boolean'
                statistics[column] = {
                    'null_count': int(series.isnull().sum())
                }
            else:
                data_type = 'string'
                
                # Calculate statistics for string columns
                non_null_series = series.dropna()
                statistics[column] = {
                    'unique_count': int(series.nunique()),
                    'max_length': int(non_null_series.astype(str).str.len().max()) if len(non_null_series) > 0 else 0,
                    'null_count': int(series.isnull().sum())
                }
            
            types[column] = data_type
            columns.append({
                'name': column,
                'type': data_type,
                'nullable': bool(series.isnull().any()),  # Convert numpy.bool_ to Python bool
                'statistics': statistics[column]
            })
        
        return {
            'columns': columns,
            'types': types,
            'row_count': int(len(df)),  # Convert numpy.int64 to Python int
            'inferred_at': datetime.now().isoformat()
        }

    def _parse_database_uri(self, uri: str) -> Dict[str, Any]:
        """Parse database URI. Uses last @ to split userinfo from host so passwords containing @ or : are correct."""
        try:
            from urllib.parse import unquote, parse_qs

            if '://' not in uri:
                raise ValueError("URI must contain ://")
            scheme_part, rest = uri.split('://', 1)
            dialect = scheme_part.split('+')[0] if '+' in scheme_part else scheme_part

            authority_end = len(rest)
            for sep in ('/', '?'):
                idx = rest.find(sep)
                if idx >= 0:
                    authority_end = min(authority_end, idx)
            authority = rest[:authority_end]
            path_and_query = rest[authority_end:].lstrip('/')
            path = path_and_query.split('?')[0] if path_and_query else ''
            query = path_and_query.split('?', 1)[1] if '?' in path_and_query else ''

            at_parts = authority.split('@')
            if len(at_parts) < 2:
                host_port = authority
                username, password = None, None
                port_idx = host_port.rfind(':')
                if port_idx >= 0:
                    host, port_str = host_port[:port_idx], host_port[port_idx + 1:]
                    try:
                        port = int(port_str)
                    except ValueError:
                        port = None
                else:
                    host, port = host_port, None
            else:
                host_port = at_parts[-1]
                userinfo = '@'.join(at_parts[:-1])
                colon_idx = userinfo.find(':')
                username = unquote(userinfo[:colon_idx]) if colon_idx >= 0 else unquote(userinfo)
                password = unquote(userinfo[colon_idx + 1:]) if colon_idx >= 0 else None
                port_idx = host_port.rfind(':')
                if port_idx >= 0:
                    host, port_str = host_port[:port_idx], host_port[port_idx + 1:]
                    try:
                        port = int(port_str)
                    except ValueError:
                        port = None
                else:
                    host, port = host_port, None

            protocol_map = {
                'postgres': 'postgresql',
                'postgresql': 'postgresql',
                'mysql': 'mysql',
                'sqlserver': 'sqlserver',
                'mssql': 'sqlserver',
            }
            db_type = protocol_map.get(dialect, dialect)
            default_ports = {'postgresql': 5432, 'mysql': 3306, 'clickhouse': 8123, 'sqlserver': 1433}
            port = port if port is not None else default_ports.get(db_type)

            ssl_mode = 'prefer'
            if query:
                qs = parse_qs(query)
                if qs.get('sslmode'):
                    ssl_mode = qs['sslmode'][0] if isinstance(qs['sslmode'], list) else str(qs['sslmode'])

            result = {
                'type': db_type,
                'host': host or None,
                'port': port,
                'database': path or '',
                'username': username,
                'password': password,
                'uri': uri,
            }
            if ssl_mode:
                result['ssl_mode'] = ssl_mode
            return result
        except Exception as e:
            raise ValueError(f"Invalid database URI format: {str(e)}")

    def _parse_odbc_connection_string(self, s: str) -> Dict[str, Any]:
        """Parse ODBC-style connection string (e.g. Server=host,port;Database=db;UID=user;PWD=pass)."""
        s = s.strip()
        if not s or ('://' in s and not ('Server=' in s or 'DRIVER=' in s)):
            return {}
        parts = {}
        for item in s.split(';'):
            item = item.strip()
            if not item or '=' not in item:
                continue
            k, v = item.split('=', 1)
            k = k.strip().lower()
            v = v.strip()
            if k == 'server':
                parts['server'] = v
            elif k == 'database' or k == 'initial catalog':
                parts['database'] = v
            elif k in ('uid', 'user id', 'username'):
                parts['username'] = v
            elif k in ('pwd', 'password'):
                parts['password'] = v
            elif k == 'trustservercertificate' and v.lower() in ('yes', 'true', '1'):
                parts['trust_server_certificate'] = True
            elif k == 'connection timeout':
                try:
                    parts['connection_timeout'] = int(v)
                except ValueError:
                    pass
        if not parts.get('server') or not parts.get('database'):
            return {}
        # Parse Server=host,port
        server = parts['server']
        host, port = server, 1433
        if ',' in server:
            host, port_str = server.rsplit(',', 1)
            try:
                port = int(port_str.strip())
            except ValueError:
                port = 1433
        return {
            'type': 'sqlserver',
            'host': host.strip(),
            'port': port,
            'database': parts.get('database', ''),
            'username': parts.get('username', ''),
            'password': parts.get('password', ''),
            'trust_server_certificate': parts.get('trust_server_certificate', True),
            'connection_timeout': parts.get('connection_timeout', 30),
        }

    def _parse_database_uri_or_odbc(self, uri_or_odbc: str) -> Dict[str, Any]:
        """Parse either a database URL (mssql+pyodbc://...) or ODBC connection string (Server=...;Database=...)."""
        s = uri_or_odbc.strip()
        # ODBC-style: contains Server= or DRIVER= and key=value pairs
        if ('Server=' in s or 'DRIVER=' in s) and (
            'Database=' in s or 'Initial Catalog=' in s or 'DATABASE=' in s
        ):
            parsed = self._parse_odbc_connection_string(s)
            if parsed:
                return parsed
        # URL-style
        if '://' in s:
            return self._parse_database_uri(s)
        return {}

    async def create_database_connection(self, config: Dict[str, Any]) -> Dict[str, Any]:
        """Create database connection using Cube.js connectors"""
        try:
            # Handle URI-based connections
            if config.get('uri'):
                logger.info(f"🔌 Parsing database URI connection")
                parsed_config = self._parse_database_uri(config['uri'])
                # Merge with original config, keeping name if provided
                config = {**parsed_config, **{k: v for k, v in config.items() if k != 'uri'}}
                logger.info(f"🔌 Parsed URI to: {config.get('type')} at {config.get('host')}:{config.get('port')}")
            
            logger.info(f"🔌 Creating database connection: {config.get('type')}")
            
            db_type = config.get('type')
            supported_dbs = self.database_connector.get_supported_databases()
            
            if db_type not in supported_dbs:
                raise ValueError(f"Unsupported database type: {db_type}. Supported: {supported_dbs}")
            
            # Create connection using DatabaseConnectorService
            cube_result = await self.database_connector.create_connection(config)
            
            if not cube_result['success']:
                raise Exception(cube_result['error'])
            
            cube_data_source = cube_result['data_source']
            
            # Generate our data source metadata
            data_source = {
                'id': cube_data_source['id'],
                'name': cube_data_source['name'],
                'type': 'database',
                'db_type': cube_data_source['db_type'],
                'host': config.get('host'),
                'database': config.get('database'),
                'created_at': datetime.now().isoformat(),
                'cube_integration': True,
                'driver': cube_data_source['driver'],
                'status': cube_data_source['status']
            }
            
            # Store in registry AND database
            self.data_sources[data_source['id']] = {
                **data_source,
                'cube_data_source': cube_data_source
            }
            
            # Also save to database for persistence
            try:
                from src.modules.data.models import DataSource as DataSourceModel
                from src.db.session import get_db
                
                # Create data source model instance
                db_data_source = DataSourceModel(
                    id=data_source['id'],
                    name=data_source['name'],
                    type='database',
                    db_type=data_source.get('db_type'),
                    connection_config=json.dumps(config),
                    metadata=json.dumps(data_source),
                    is_active=True
                )
                
                # Add to session and commit
                db = get_db()
                await db.add(db_data_source)
                logger.info(f"✅ Data source saved to database: {data_source['id']}")
            except Exception as db_error:
                logger.warning(f"⚠️ Could not save to database: {db_error}")
                # Continue anyway since we have in-memory storage
            
            logger.info(f"✅ Cube.js database connection created successfully: {data_source['id']}")
            
            return {
                'success': True,
                'data_source': {
                    'id': data_source['id'],
                    'name': data_source['name'],
                    'type': data_source['type'],
                    'db_type': data_source['db_type'],
                    'created_at': data_source['created_at'],
                    'cube_integration': True,
                    'driver': data_source['driver'],
                    'status': data_source['status']
                }
            }
            
        except Exception as error:
            logger.error(f"❌ Cube.js database connection failed: {str(error)}")
            return {
                'success': False,
                'error': str(error)
            }

    async def query_data_source(
        self,
        data_source_id: str,
        query: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """Query data from data source"""
        try:
            data_source = self.data_sources.get(data_source_id)
            if not data_source:
                raise ValueError(f"Data source not found: {data_source_id}")
            
            if query is None:
                query = {}
            
            if data_source['type'] == 'file':
                return await self._query_file_data_source(data_source, query)
            elif data_source['type'] == 'database':
                return await self._query_database_data_source(data_source, query)
            else:
                raise ValueError(f"Unsupported data source type: {data_source['type']}")
                
        except Exception as error:
            logger.error(f"❌ Data source query failed: {str(error)}")
            return {
                'success': False,
                'error': str(error)
            }

    async def _query_file_data_source(
        self, 
        data_source: Dict[str, Any], 
        query: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Query file-based data source"""
        # CRITICAL: Check if data is in memory first
        data = data_source.get('data', [])
        
        # If no in-memory data, try to load from edition-specific object storage.
        if not data or len(data) == 0:
            object_key = data_source.get('file_path')  # Now it's object_key
            if object_key:
                try:
                    storage_service = UploadDatasourceStorageService()
                    project_id = data_source.get('project_id')
                    
                    if not project_id:
                        logger.warning("⚠️ project_id not found in data_source, cannot load from datasource storage")
                    else:
                        logger.info(
                            "📊 Loading file data from %s storage: %s",
                            storage_service.storage_type,
                            object_key,
                        )
                        file_content = await storage_service.get_file(object_key, project_id)
                        
                        # Write to temp file for processing
                        file_format = data_source.get('format', 'csv')
                        schema_obj = data_source.get('schema') if isinstance(data_source.get('schema'), dict) else {}
                        storage_format = ((schema_obj.get('storage') or {}).get('format') if isinstance(schema_obj, dict) else None)
                        blob_format = (storage_format or file_format or 'csv').lower()
                        with tempfile.NamedTemporaryFile(delete=False, suffix=f".{blob_format}") as tmp:
                            tmp.write(file_content)
                            tmp_path = tmp.name
                        
                        try:
                            # Process file
                            data = await self._read_file_data(tmp_path, blob_format, limit=query.get('limit', DEFAULT_FILE_QUERY_LIMIT))
                            logger.info("✅ Loaded %d rows from datasource storage", len(data))
                        finally:
                            # Clean up temp file
                            if os.path.exists(tmp_path):
                                os.unlink(tmp_path)
                except Exception as e:
                    logger.error(f"❌ Failed to load from datasource storage: {str(e)}")
                    data = []
        
        # Fallback to sample_data
        if not data or len(data) == 0:
            data = data_source.get('sample_data', []) or data_source.get('preview_data', [])
            if data:
                logger.info(f"📊 Using sample/preview data ({len(data)} rows)")
        
        # Apply filters
        filters = query.get('filters', [])
        for filter_item in filters:
            column = filter_item.get('column')
            operator = filter_item.get('operator')
            value = filter_item.get('value')
            
            if operator == 'equals':
                data = [row for row in data if str(row.get(column, '')).lower() == str(value).lower()]
            elif operator == 'contains':
                data = [row for row in data if str(value).lower() in str(row.get(column, '')).lower()]
            elif operator == 'greater_than':
                data = [row for row in data if float(row.get(column, 0) or 0) > float(value)]
            elif operator == 'less_than':
                data = [row for row in data if float(row.get(column, 0) or 0) < float(value)]
        
        # Apply sorting
        sort_config = query.get('sort')
        if sort_config:
            column = sort_config.get('column')
            direction = sort_config.get('direction', 'asc')
            reverse = direction == 'desc'
            
            try:
                data = sorted(data, key=lambda x: x.get(column, ''), reverse=reverse)
            except:
                # If sorting fails, continue without sorting
                pass
        
        # Apply pagination
        offset = query.get('offset', 0)
        limit = query.get('limit', DEFAULT_PAGE_LIMIT)
        total_rows = len(data)
        paginated_data = data[offset:offset + limit]
        
        return {
            'success': True,
            'data': paginated_data,
            'total_rows': total_rows,
            'offset': offset,
            'limit': limit,
            'schema': data_source.get('schema')
        }

    async def _query_database_data_source(
        self,
        data_source: Dict[str, Any],
        query: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Query database data source (NoSQL connectors or Cube.js)."""
        db_type = (data_source.get('db_type') or data_source.get('format') or '').lower()
        config = data_source.get('config') or data_source.get('connection_config') or {}

        if db_type == 'mongodb' and MONGODB_AVAILABLE and MongoDBConnector:
            try:
                conn_str = config.get('connection_string') or config.get('host') or ''
                database = config.get('database') or ''
                if not conn_str or not database:
                    return {"success": False, "error": "MongoDB config missing connection_string and database."}
                connector = MongoDBConnector(conn_str, database)
                await asyncio.to_thread(connector.connect)
                nl_query = (query.get('query') or query.get('sql') or query.get('nl_query') or '').strip() or str(query)
                collection = (query.get('collection') or config.get('collection') or '').strip()
                if not collection and config.get('database'):
                    collection = list(connector.db.list_collection_names())[:1][0] if connector.db else ''
                if not collection:
                    connector.disconnect()
                    return {"success": False, "error": "MongoDB query requires collection name."}
                result = await connector.execute_nl_query(nl_query, collection, limit=query.get('limit', 1000))
                connector.disconnect()
                return {"success": True, "data": result.get('data', []), "row_count": len(result.get('data', []))}
            except Exception as e:
                logger.exception("MongoDB query failed")
                return {"success": False, "error": str(e)}
        if db_type == 'cassandra' and CASSANDRA_AVAILABLE and CassandraConnector:
            try:
                host_str = config.get('host') or config.get('contact_points') or ''
                contact_points = [x.strip() for x in host_str.split(',') if x.strip()] if isinstance(host_str, str) else host_str
                keyspace = config.get('keyspace') or config.get('database') or ''
                if not contact_points or not keyspace:
                    return {"success": False, "error": "Cassandra config missing host and keyspace."}
                connector = CassandraConnector(contact_points, keyspace, config.get('username'), config.get('password'))
                await asyncio.to_thread(connector.connect)
                nl_query = (query.get('query') or query.get('sql') or query.get('cql') or '').strip() or str(query)
                table = (query.get('table') or config.get('table') or '').strip()
                if not table:
                    connector.disconnect()
                    return {"success": False, "error": "Cassandra query requires table name."}
                result = await connector.execute_nl_query(nl_query, table, limit=query.get('limit', 1000))
                connector.disconnect()
                return {"success": True, "data": result.get('data', []), "row_count": len(result.get('data', []))}
            except Exception as e:
                logger.exception("Cassandra query failed")
                return {"success": False, "error": str(e)}
        if db_type == 'dynamodb' and DYNAMODB_AVAILABLE and DynamoDBConnector:
            try:
                region = config.get('region') or 'us-east-1'
                access_key = config.get('access_key_id') or config.get('accessKey')
                secret_key = config.get('secret_access_key') or config.get('secretKey')
                if not access_key or not secret_key:
                    return {"success": False, "error": "DynamoDB config missing credentials."}
                connector = DynamoDBConnector(region, access_key, secret_key)
                table_name = (query.get('table') or config.get('table_name') or config.get('database') or '').strip()
                if not table_name:
                    return {"success": False, "error": "DynamoDB query requires table name."}
                nl_query = (query.get('query') or query.get('sql') or '').strip() or str(query)
                result = await connector.execute_nl_query(nl_query, table_name, limit=query.get('limit', 1000))
                return {"success": True, "data": result.get('data', []), "row_count": len(result.get('data', []))}
            except Exception as e:
                logger.exception("DynamoDB query failed")
                return {"success": False, "error": str(e)}

        # Fall through to Cube.js / SQL handling
        try:
            # If not using Cube.js, attempt direct SQL execution using MultiEngineQueryService
            if not data_source.get('cube_integration'):
                from src.modules.data.services.multi_engine_query_service import MultiEngineQueryService, get_multi_engine_query_service
                multi = get_multi_engine_query_service()
                
                # Build SQL if missing but table and columns are provided in query
                sql_query = query.get('sql') or query.get('query')
                if not sql_query and isinstance(query, dict):
                    table = query.get('table')
                    if not table:
                        schema_info = data_source.get('schema', {})
                        if isinstance(schema_info, str): schema_info = json.loads(schema_info)
                        table = schema_info.get('table') or (schema_info.get('tables')[0].get('name') if schema_info.get('tables') else None)
                    
                    if table:
                        limit = query.get('limit', 1000)
                        sql_query = f"SELECT * FROM {table} LIMIT {limit}"
                
                if not sql_query:
                    return {
                        'success': False,
                        'error': 'Database query requires SQL string or table name'
                    }
                
                result = await multi.execute_query(sql_query, data_source)
                return result
            
            cube_data_source = data_source.get('cube_data_source')
            if not cube_data_source:
                return {
                    'success': False,
                    'error': 'Cube.js data source not found'
                }
            
            # Execute query through DatabaseConnectorService
            # Note: query should be a SQL string, not a structured query
            sql_query = query.get('sql') or query.get('query') if isinstance(query, dict) else str(query)
            
            result = await self.database_connector.execute_query(
                data_source['id'],
                sql_query
            )
            
            return result
            
        except Exception as error:
            logger.error(f"❌ Database query failed: {str(error)}")
            return {
                'success': False,
                'error': str(error)
            }

    async def get_data_source(self, data_source_id: str) -> Dict[str, Any]:
        """Get a specific data source by ID - checks both in-memory and database"""
        try:
            # First check in-memory cache
            data_source = self.data_sources.get(data_source_id)
            if data_source:
                logger.info(f"✅ Found data source {data_source_id} in memory cache")
                return {
                    'success': True,
                    'data_source': data_source
                }
            
            # If not in memory, check database
            logger.info(f"🔍 Data source {data_source_id} not in memory, checking database...")
            from src.modules.data.models import DataSource
            from src.db.session import async_session
            import json

            async with async_session() as db:
                from sqlalchemy import select
                
                query = select(DataSource).where(
                    DataSource.id == data_source_id,
                    DataSource.is_active == True
                )
                result = await db.execute(query)
                db_source = result.scalar_one_or_none()
                
                if db_source:
                    logger.info(f"✅ Found data source {data_source_id} in database")
                    # Convert database model to dict format
                    data_source_dict = {
                        'id': str(db_source.id),
                        'name': db_source.name,
                        'type': db_source.type,
                        'format': db_source.format,
                        'description': db_source.description,
                        'project_id': str(db_source.project_id) if db_source.project_id else None,
                        'row_count': db_source.row_count,
                        'size': db_source.size,
                        'is_active': db_source.is_active,
                        'created_at': db_source.created_at.isoformat() if db_source.created_at else None,
                        'updated_at': db_source.updated_at.isoformat() if db_source.updated_at else None,
                    }
                    
                    # Add file_path if available directly from model (now stores object_key)
                    if hasattr(db_source, 'file_path') and db_source.file_path:
                        data_source_dict['file_path'] = db_source.file_path  # This is now object_key
                    
                    # Parse config and schema if available
                    try:
                        if db_source.connection_config:
                            config = json.loads(db_source.connection_config) if isinstance(db_source.connection_config, str) else db_source.connection_config
                            try:
                                from src.modules.data.utils.credentials import decrypt_credentials
                                config = decrypt_credentials(config)
                            except Exception:
                                pass
                            data_source_dict['config'] = config
                            data_source_dict['connection_config'] = config
                    except Exception as e:
                        logger.warning(f"⚠️ Failed to parse connection_config: {e}")
                    
                    try:
                        if db_source.schema:
                            schema = json.loads(db_source.schema) if isinstance(db_source.schema, str) else db_source.schema
                            data_source_dict['schema'] = schema
                    except Exception as e:
                        logger.warning(f"⚠️ Failed to parse schema: {e}")
                    
                    # For sample_duckdb: if schema empty, fetch from shared DuckDB
                    if db_source.type == 'sample_duckdb':
                        stored = data_source_dict.get('schema')
                        has_tables = isinstance(stored, dict) and bool(stored.get('tables'))
                        if not has_tables:
                            try:
                                schema_result = await self.get_sample_duckdb_schema(data_source_dict)
                                if schema_result.get('success') and schema_result.get('schema'):
                                    data_source_dict['schema'] = schema_result['schema']
                                    if schema_result.get('data_source', {}).get('row_count') is not None:
                                        data_source_dict['row_count'] = schema_result['data_source']['row_count']
                            except Exception as schema_err:
                                logger.warning(f"⚠️ Could not load sample_duckdb schema: {schema_err}")
                    
                    # For google_sheets: if schema empty, fetch from sheet URL (CSV export)
                    if db_source.type == 'google_sheets':
                        stored = data_source_dict.get('schema')
                        has_tables = isinstance(stored, dict) and bool(stored.get('tables'))
                        if not has_tables:
                            try:
                                schema_result = await self.get_google_sheets_schema(data_source_dict)
                                if schema_result.get('success') and schema_result.get('schema'):
                                    data_source_dict['schema'] = schema_result['schema']
                                    if schema_result.get('data_source', {}).get('row_count') is not None:
                                        data_source_dict['row_count'] = schema_result['data_source']['row_count']
                                    logger.info(f"✅ Loaded google_sheets schema for {data_source_id}")
                            except Exception as schema_err:
                                logger.warning(f"⚠️ Could not load google_sheets schema for {data_source_id}: {schema_err}")
                    
                    # Load sample_data from database
                    if db_source.sample_data:
                        try:
                            sample = json.loads(db_source.sample_data) if isinstance(db_source.sample_data, str) else db_source.sample_data
                            data_source_dict['sample_data'] = sample
                            data_source_dict['data'] = sample  # For compatibility
                            logger.info(f"✅ Loaded sample_data from database")
                        except Exception as e:
                            logger.warning(f"⚠️ Failed to parse sample_data: {e}")
                    
                    # Add original_filename if available
                    if hasattr(db_source, 'original_filename') and db_source.original_filename:
                        data_source_dict['original_filename'] = db_source.original_filename
                    
                    # Cache in memory for future access
                    self.data_sources[data_source_id] = data_source_dict
                    
                    return {
                        'success': True,
                        'data_source': data_source_dict
                    }
                else:
                    logger.warning(f"⚠️ Data source {data_source_id} not found in database or memory")
                    return {
                        'success': False,
                        'error': 'Data source not found'
                    }
        except Exception as e:
            logger.error(f"❌ Failed to get data source: {str(e)}", exc_info=True)
            return {
                'success': False,
                'error': str(e)
            }

    def list_data_sources(self) -> Dict[str, Any]:
        """List all data sources"""
        sources = []
        for source in self.data_sources.values():
            metadata = {key: value for key, value in source.items() 
                       if key not in ['data', 'connection']}
            sources.append(metadata)
        
        return {
            'success': True,
            'data_sources': sources,
            'count': len(sources)
        }

    def delete_data_source(self, data_source_id: str) -> Dict[str, Any]:
        """Delete data source"""
        try:
            data_source = self.data_sources.get(data_source_id)
            
            # Attempt DB deletion first (if persisted)
            try:
                from src.modules.data.models import DataSource
                from src.modules.data.models import ProjectDataSource
                # Use async_session factory directly (get_async_session is an async generator for FastAPI deps)
                from src.db.session import async_session
                import asyncio
                async def _delete_from_db():
                    async with async_session() as db:
                        from sqlalchemy import delete, select
                        
                        # First, fetch the data source to get file_path/project_id for object storage deletion
                        try:
                            result = await db.execute(
                                select(DataSource).where(DataSource.id == data_source_id)
                            )
                            db_source = result.scalar_one_or_none()
                            
                            # Delete stored file if it's a file source
                            if db_source:
                                ds_type = getattr(db_source, 'type', None)
                                ds_file_path = getattr(db_source, 'file_path', None)
                                
                                if ds_type == 'file' and ds_file_path:
                                    file_path = str(ds_file_path)
                                    # Check if file_path is an object_key (new or legacy paths)
                                    is_blob_key = (
                                        file_path.startswith("orgs/")
                                        or file_path.startswith("projects/")
                                        or file_path.startswith("org_files/")
                                        or file_path.startswith("project_files/")
                                        or file_path.startswith("user_files/")
                                    )
                                    if is_blob_key:
                                        try:
                                            storage_service = UploadDatasourceStorageService()
                                            project_id_str = str(db_source.project_id) if db_source.project_id else ""
                                            await storage_service.delete_file(file_path, project_id_str)
                                            logger.info(f"✅ Deleted file from datasource storage during data source deletion: {file_path}")
                                        except Exception as e:
                                            logger.warning(f"⚠️ Failed to delete file from datasource storage for data_source {data_source_id}: {e}")
                        except Exception as e:
                            logger.warning(f"⚠️ Failed to fetch data source for file deletion: {e}")
                        
                        # Remove project links
                        try:
                            await db.execute(delete(ProjectDataSource).where(ProjectDataSource.data_source_id == data_source_id))
                        except Exception:
                            pass
                        # Remove data source row
                        await db.execute(delete(DataSource).where(DataSource.id == data_source_id))
                        await db.commit()
                # Run async deletion in current loop if available
                try:
                    loop = asyncio.get_event_loop()
                    if loop.is_running():
                        fut = asyncio.ensure_future(_delete_from_db())
                        # fire-and-forget
                    else:
                        loop.run_until_complete(_delete_from_db())
                except RuntimeError:
                    # No loop, run new loop
                    asyncio.run(_delete_from_db())
            except Exception as db_del_err:
                logger.warning(f"DB deletion for data source {data_source_id} skipped/failed: {db_del_err}")

            # In-memory cleanup
            if data_source:
                # Clean up file if it's a file-based source
                if data_source.get('type') == 'file' and 'file_path' in data_source:
                    file_path = data_source['file_path']
                    user_id = data_source.get('user_id')
                    
                    # Check if file_path is an object_key (new or legacy paths)
                    is_blob_key = file_path and (
                        file_path.startswith("orgs/")
                        or file_path.startswith("projects/")
                        or file_path.startswith("org_files/")
                        or file_path.startswith("project_files/")
                        or file_path.startswith("user_files/")
                    )
                    if is_blob_key:
                        project_id = data_source.get('project_id') or user_id
                        if project_id:
                            try:
                                import asyncio
                                storage_service = UploadDatasourceStorageService()
                                
                                async def _delete_from_storage():
                                    await storage_service.delete_file(file_path, str(project_id))
                                
                                # Run async deletion
                                try:
                                    loop = asyncio.get_event_loop()
                                    if loop.is_running():
                                        asyncio.ensure_future(_delete_from_storage())
                                    else:
                                        loop.run_until_complete(_delete_from_storage())
                                except RuntimeError:
                                    asyncio.run(_delete_from_storage())
                                logger.info(f"✅ Deleted file from datasource storage (in-memory cleanup): {file_path}")
                            except Exception as e:
                                logger.warning(f"⚠️ Failed to delete file from datasource storage (in-memory): {e}")
                        else:
                            logger.warning(f"⚠️ Cannot delete file from datasource storage: project_id missing in data_source {data_source_id}")
                    elif file_path and os.path.exists(file_path):
                        # Legacy: local file path (shouldn't happen with new uploads, but handle for backwards compatibility)
                        try:
                            os.unlink(file_path)
                            logger.info(f"✅ Deleted local file: {file_path}")
                        except Exception as e:
                            logger.warning(f"⚠️ Failed to delete local file: {e}")
                # Close database connection if tracked
                if data_source.get('type') == 'database' and 'connection' in data_source:
                    connection = data_source['connection']
                    if hasattr(connection, 'close'):
                        try:
                            connection.close()
                        except Exception:
                            pass
                self.data_sources.pop(data_source_id, None)

            from src.modules.data.services.pool_invalidation import dispose_direct_sql_pool_for_data_source
            dispose_direct_sql_pool_for_data_source(data_source_id)

            return {'success': True, 'message': 'Data source deleted successfully'}
        except Exception as e:
            logger.error(f"Failed to delete data source {data_source_id}: {e}")
            return {'success': False, 'error': str(e)}

    async def generate_data_insights(self, data_source_id: str) -> Dict[str, Any]:
        """Generate AI insights for a data source"""
        try:
            logger.info(f"🔍 Generating AI insights for data source: {data_source_id}")
            
            # Get the data source
            data_source = self.data_sources.get(data_source_id)
            if not data_source:
                return {'success': False, 'error': 'Data source not found'}
            
            # Get data and schema
            data = data_source.get('data', [])
            schema = data_source.get('schema', {})
            name = data_source.get('name', 'Unknown')
            
            # Generate insights using AI
            insights_result = await self.ai_schema_service.generate_data_insights(
                data, schema, name
            )
            
            return insights_result
            
        except Exception as e:
            logger.error(f"❌ Failed to generate insights: {str(e)}")
            return {'success': False, 'error': str(e)}

    async def _fetch_api_schema(
        self, data_source_id: str, data_source: Any
    ) -> Dict[str, Any]:
        """Fetch from API endpoint and infer schema from JSON/CSV response. Supports base_url + path, method, headers, auth."""
        try:
            config = _parse_json_field(data_source.connection_config)
            if not config:
                return {"success": False, "error": "No connection configuration for this API source. Edit the data source and set Base URL (and optional Path)."}
            try:
                from src.modules.data.utils.credentials import decrypt_credentials
                config = decrypt_credentials(dict(config))
            except Exception:
                config = dict(config) if config else {}
            base = (config.get("url") or config.get("base_url") or "").strip().rstrip("/")
            path = (config.get("path") or "").strip().lstrip("/")
            api_url = f"{base}/{path}" if path else base
            if not api_url:
                return {"success": False, "error": "API URL is required. Set Base URL in the data source connection."}
            method = (config.get("method") or "GET").upper()
            headers = config.get("headers") or {}
            if isinstance(headers, str):
                try:
                    headers = json.loads(headers)
                except json.JSONDecodeError:
                    headers = {}
            params = config.get("params") or {}
            auth = None
            auth_type = (config.get("auth_type") or "").strip().lower() or "api_key"
            if auth_type == "basic" and config.get("username") is not None and config.get("password") is not None:
                try:
                    auth = aiohttp.BasicAuth(str(config["username"]), str(config["password"]))
                except Exception:
                    pass
            elif auth_type == "bearer" and config.get("bearer_token"):
                headers["Authorization"] = f"Bearer {config['bearer_token']}"
            elif auth_type == "api_key" and config.get("api_key"):
                headers[config.get("api_key_header", "X-API-Key")] = config["api_key"]
            elif auth_type != "none":
                if config.get("api_key"):
                    headers[config.get("api_key_header", "X-API-Key")] = config["api_key"]
                elif config.get("bearer_token"):
                    headers["Authorization"] = f"Bearer {config['bearer_token']}"
                elif config.get("username") and config.get("password"):
                    try:
                        auth = aiohttp.BasicAuth(str(config["username"]), str(config["password"]))
                    except Exception:
                        pass
            import aiohttp
            async with aiohttp.ClientSession() as session:
                _timeout = max(5, min(120, int(config.get("timeout", 30))))
                async with session.request(
                    method, api_url, headers=headers, params=params, auth=auth,
                    timeout=aiohttp.ClientTimeout(total=_timeout)
                ) as response:
                    if response.status != 200:
                        return {
                            "success": False,
                            "error": f"API returned {response.status}. Check Base URL, Path, and authentication.",
                        }
                    content_type = (response.headers.get("Content-Type") or "").lower()
                    text = await response.text()
            rows_for_schema: List[Dict[str, Any]] = []
            if "json" in content_type:
                try:
                    data = json.loads(text)
                except json.JSONDecodeError:
                    return {"success": False, "error": "API response is not valid JSON."}
                if isinstance(data, list) and data:
                    rows_for_schema = data[:100]
                elif isinstance(data, dict):
                    for key in ("data", "results", "items", "records"):
                        if key in data and isinstance(data[key], list) and data[key]:
                            rows_for_schema = data[key][:100]
                            break
                    if not rows_for_schema and data:
                        rows_for_schema = [data]
            elif "csv" in content_type or not rows_for_schema:
                try:
                    import io
                    df = pd.read_csv(io.StringIO(text), nrows=100)
                    rows_for_schema = df.to_dict("records")
                except Exception:
                    try:
                        data = json.loads(text)
                        if isinstance(data, list) and data:
                            rows_for_schema = data[:100]
                    except Exception:
                        pass
            if not rows_for_schema or not isinstance(rows_for_schema[0], dict):
                return {"success": False, "error": "Could not infer schema from API response. Ensure the endpoint returns JSON array or object with 'data'/'results'/'items', or CSV."}
            first = rows_for_schema[0]
            columns = []
            for col_name, col_value in first.items():
                col_type = "string"
                if col_value is not None:
                    if isinstance(col_value, bool):
                        col_type = "boolean"
                    elif isinstance(col_value, int):
                        col_type = "number"
                    elif isinstance(col_value, float):
                        col_type = "number"
                    elif isinstance(col_value, str) and len(col_value) >= 8:
                        try:
                            datetime.fromisoformat(col_value.replace("Z", "+00:00")[:26])
                            col_type = "date"
                        except Exception:
                            pass
                columns.append({"name": col_name, "type": col_type, "nullable": True})
            schema = {
                "tables": [{"name": "data", "columns": columns, "row_count": len(rows_for_schema)}],
                "connection_database": "default",
                "last_updated": datetime.now().isoformat(),
            }
            logger.info("API schema inferred: %d columns from %s", len(columns), data_source_id)
            return {"success": True, "schema": schema, "row_count": len(rows_for_schema)}
        except Exception as e:
            logger.warning("_fetch_api_schema failed: %s", e)
            return {"success": False, "error": str(e)}

    async def get_sample_duckdb_schema(self, data_source: Any) -> Dict[str, Any]:
        """Get schema for a sample_duckdb data source from the shared DuckDB file.
        Returns same shape as get_database_schema: { success, schema: { tables, schemas }, data_source }.
        data_source can be an ORM object or a dict with connection_config, id, name.
        Schema is cached per domain (10 min TTL) so many users/orgs/projects share one file without repeated I/O.
        """
        global _SAMPLE_DUCKDB_SCHEMA_CACHE
        try:
            import duckdb
            config = (data_source.get("connection_config") if isinstance(data_source, dict) else getattr(data_source, "connection_config", None))
            if isinstance(config, str):
                try:
                    config = json.loads(config)
                except json.JSONDecodeError:
                    config = {}
            if not isinstance(config, dict):
                config = {}
            try:
                config = decrypt_credentials(config)
            except Exception as dec_err:
                logger.debug("Decrypt connection_config for sample_duckdb: %s", dec_err)
            domain = (config.get("domain") or "banking").strip().lower()
            _id = data_source.get("id") if isinstance(data_source, dict) else getattr(data_source, "id", None)
            _name = data_source.get("name", "") if isinstance(data_source, dict) else getattr(data_source, "name", "")
            # Use per-domain cache so many tenants don't hit the file on every schema request
            now = time.time()
            if domain in _SAMPLE_DUCKDB_SCHEMA_CACHE:
                entry = _SAMPLE_DUCKDB_SCHEMA_CACHE[domain]
                if (now - entry.get("ts", 0)) < _SAMPLE_DUCKDB_SCHEMA_CACHE_TTL_SEC and entry.get("schema"):
                    return {
                        "success": True,
                        "schema": entry["schema"],
                        "data_source": {
                            "id": _id,
                            "name": _name,
                            "type": "sample_duckdb",
                            "row_count": entry.get("row_count", 0),
                        },
                    }
            sample_path = os.getenv("SAMPLE_DATA_DUCKDB_PATH", "").strip()
            if not sample_path or not os.path.isfile(sample_path):
                sample_path = "/app/scripts/sample-data/duckdb/sample_data.duckdb"
            if not sample_path or not os.path.isfile(sample_path):
                _base = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..", "..", ".."))
                sample_path = os.path.join(_base, "scripts", "sample-data", "duckdb", "sample_data.duckdb")
            if not sample_path or not os.path.isfile(sample_path):
                sample_path = "/app/scripts/sample-data/duckdb/sample_data.duckdb"
            if not sample_path or not os.path.isfile(sample_path):
                logger.debug("Sample DuckDB file not found at %s (returning empty schema)", sample_path or "SAMPLE_DATA_DUCKDB_PATH")
                return {
                    "success": False,
                    "error": "Sample data DuckDB file not found. Set SAMPLE_DATA_DUCKDB_PATH or run sample data generators.",
                    "schema": {"tables": [], "schemas": []},
                    "data_source": {
                        "id": data_source.get("id") if isinstance(data_source, dict) else getattr(data_source, "id", None),
                        "name": data_source.get("name", "") if isinstance(data_source, dict) else getattr(data_source, "name", ""),
                        "type": "sample_duckdb",
                        "row_count": 0,
                    },
                }
            conn = duckdb.connect(sample_path, read_only=True)
            try:
                tables_rows = conn.execute(
                    "SELECT table_schema, table_name FROM information_schema.tables "
                    "WHERE table_schema = ? AND table_schema NOT IN ('information_schema', 'pg_catalog')",
                    [domain],
                ).fetchall()
                tables = []
                total_rows = 0
                for (sch, tbl) in tables_rows:
                    cols = conn.execute(
                        "SELECT column_name, data_type, is_nullable FROM information_schema.columns "
                        "WHERE table_schema = ? AND table_name = ? ORDER BY ordinal_position",
                        [sch, tbl],
                    ).fetchall()
                    columns = [
                        {
                            "name": c[0],
                            "type": (c[1] or "VARCHAR"),
                            "nullable": (str(c[2]).upper() == "YES") if c[2] is not None else True,
                            "primary_key": False,
                        }
                        for c in cols
                    ]
                    try:
                        cnt = conn.execute(f'SELECT COUNT(*) FROM "{sch}"."{tbl}"').fetchone()
                        row_count = int(cnt[0]) if cnt else 0
                    except Exception:
                        row_count = 0
                    total_rows += row_count
                    tables.append({
                        "schema": sch,
                        "name": tbl,
                        "columns": columns,
                        "rowCount": row_count,
                    })
                schema_obj = {
                    "tables": tables,
                    "schemas": [domain] if tables else [],
                    "last_updated": datetime.now().isoformat(),
                }
                logger.info("Sample DuckDB schema for domain %s: %s tables", domain, len(tables))
                # Cache by domain so many users/orgs/projects don't hit the file on every request
                try:
                    _SAMPLE_DUCKDB_SCHEMA_CACHE[domain] = {
                        "schema": schema_obj,
                        "row_count": total_rows,
                        "ts": time.time(),
                    }
                except Exception:
                    pass
                return {
                    "success": True,
                    "schema": schema_obj,
                    "data_source": {
                        "id": _id,
                        "name": _name,
                        "type": "sample_duckdb",
                        "row_count": total_rows,
                    },
                }
            finally:
                conn.close()
        except Exception as e:
            logger.exception("Failed to get sample_duckdb schema: %s", e)
            _id = data_source.get("id") if isinstance(data_source, dict) else getattr(data_source, "id", None)
            _name = data_source.get("name", "") if isinstance(data_source, dict) else getattr(data_source, "name", "")
            return {
                "success": False,
                "error": str(e),
                "schema": {"tables": [], "schemas": []},
                "data_source": {
                    "id": _id,
                    "name": _name,
                    "type": "sample_duckdb",
                    "row_count": 0,
                },
            }

    async def get_google_sheets_schema(self, data_source: Any) -> Dict[str, Any]:
        """Get schema for a google_sheets data source by fetching CSV export (works for public/published sheets).
        Returns same shape as get_sample_duckdb_schema: { success, schema: { tables }, data_source }.
        """
        try:
            config = (
                data_source.get("connection_config")
                if isinstance(data_source, dict)
                else getattr(data_source, "connection_config", None)
            )
            if isinstance(config, str):
                try:
                    config = json.loads(config)
                except json.JSONDecodeError:
                    config = {}
            if not isinstance(config, dict):
                config = {}
            sheet_url = (config.get("sheet_url") or "").strip()
            if not sheet_url:
                raise ValueError("google_sheets requires connection_config.sheet_url")
            sheet_id, gid = _parse_google_sheet_url(sheet_url, config.get("gid"))
            export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
            import aiohttp
            async with aiohttp.ClientSession() as session:
                async with session.get(export_url, timeout=aiohttp.ClientTimeout(total=15)) as resp:
                    if resp.status != 200:
                        text = await resp.text()
                        raise ValueError(f"Google Sheets export returned {resp.status}: {text[:200]}")
                    body = await resp.text()
            if not body or not body.strip():
                raise ValueError("Google Sheet export returned empty content")
            # Parse CSV (first row = headers)
            import csv
            from io import StringIO
            reader = csv.reader(StringIO(body))
            rows = list(reader)
            if not rows:
                raise ValueError("Sheet has no rows")
            headers = [h.strip() or f"column_{i}" for i, h in enumerate(rows[0])]
            sample_rows = rows[1:][: FILE_UPLOAD_SAMPLE_ROWS] if len(rows) > 1 else []
            columns = []
            for i, col_name in enumerate(headers):
                col_type = "string"
                for row in sample_rows:
                    if i < len(row) and row[i].strip():
                        val = row[i].strip()
                        try:
                            float(val.replace(",", ""))
                            col_type = "number"
                            break
                        except ValueError:
                            pass
                columns.append({"name": col_name, "type": col_type, "nullable": True, "primary_key": False})
            row_count = max(0, len(rows) - 1)
            schema_obj = {
                "tables": [
                    {
                        "name": "data",
                        "schema": "public",
                        "columns": columns,
                        "rowCount": row_count,
                    }
                ],
                "schemas": ["public"],
                "last_updated": datetime.now().isoformat(),
            }
            _id = data_source.get("id") if isinstance(data_source, dict) else getattr(data_source, "id", None)
            _name = data_source.get("name", "") if isinstance(data_source, dict) else getattr(data_source, "name", "")
            return {
                "success": True,
                "schema": schema_obj,
                "data_source": {
                    "id": _id,
                    "name": _name,
                    "type": "google_sheets",
                    "row_count": row_count,
                },
            }
        except Exception as e:
            logger.warning("get_google_sheets_schema failed: %s", e)
            _id = data_source.get("id") if isinstance(data_source, dict) else getattr(data_source, "id", None)
            _name = data_source.get("name", "") if isinstance(data_source, dict) else getattr(data_source, "name", "")
            return {
                "success": False,
                "error": str(e),
                "schema": {"tables": [], "schemas": []},
                "data_source": {"id": _id, "name": _name, "type": "google_sheets", "row_count": 0},
            }

    async def get_google_sheets_data(
        self, data_source: Any, limit: int = 5000
    ) -> Dict[str, Any]:
        """Fetch Google Sheet rows for data panel preview. Returns { success, data: list of dicts, row_count }."""
        try:
            config = (
                data_source.get("connection_config")
                if isinstance(data_source, dict)
                else getattr(data_source, "connection_config", None)
            )
            if isinstance(config, str):
                try:
                    config = json.loads(config)
                except json.JSONDecodeError:
                    config = {}
            if not isinstance(config, dict):
                config = {}
            sheet_url = (config.get("sheet_url") or "").strip()
            if not sheet_url:
                return {"success": False, "error": "sheet_url required", "data": [], "row_count": 0}
            sheet_id, gid = _parse_google_sheet_url(sheet_url, config.get("gid"))
            export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
            import aiohttp
            async with aiohttp.ClientSession() as session:
                async with session.get(export_url, timeout=aiohttp.ClientTimeout(total=15)) as resp:
                    if resp.status != 200:
                        return {"success": False, "error": f"Export returned {resp.status}", "data": [], "row_count": 0}
                    body = await resp.text()
            if not body or not body.strip():
                return {"success": True, "data": [], "row_count": 0}
            import csv
            from io import StringIO
            reader = csv.reader(StringIO(body))
            rows = list(reader)
            if not rows:
                return {"success": True, "data": [], "row_count": 0}
            headers = [h.strip() or f"column_{i}" for i, h in enumerate(rows[0])]
            data_rows = rows[1:][:limit]
            data = [dict(zip(headers, row[: len(headers)])) for row in data_rows]
            return {"success": True, "data": data, "row_count": len(data)}
        except Exception as e:
            logger.warning("get_google_sheets_data failed: %s", e)
            return {"success": False, "error": str(e), "data": [], "row_count": 0}

    async def get_database_schema(self, data_source_id: str, force_refresh: bool = False) -> Dict[str, Any]:
        """Get database schema information for a connected database"""
        try:
            logger.info(f"🔍 Fetching database schema for: {data_source_id}")
            
            # Get the data source from database
            from src.modules.data.models import DataSource
            from src.db.session import async_session

            async with async_session() as db:
                from sqlalchemy import select
                
                query = select(DataSource).where(DataSource.id == data_source_id)
                result = await db.execute(query)
                data_source = result.scalar_one_or_none()
                
                if not data_source:
                    return {
                        'success': False,
                        'error': 'Data source not found'
                    }
                # API sources: fetch from endpoint and infer schema from response (JSON/CSV)
                if data_source.type == 'api':
                    api_result = await self._fetch_api_schema(data_source_id, data_source)
                    if api_result.get('success') and api_result.get('schema'):
                        return {
                            'success': True,
                            'schema': api_result['schema'],
                            'data_source': {
                                'id': data_source.id,
                                'name': data_source.name,
                                'type': data_source.type,
                                'row_count': api_result.get('row_count'),
                            },
                        }
                    return {
                        'success': False,
                        'error': api_result.get('error', 'Failed to fetch API schema'),
                    }
                if data_source.type != 'database' and data_source.type != 'warehouse':
                    return {
                        'success': False,
                        'error': 'Data source is not a database or warehouse connection'
                    }
                
                # Parse the stored schema/config (DB JSON column may return dict or str)
                config = _parse_json_field(data_source.connection_config)
                schema_info = _parse_json_field(data_source.schema)
                if not config:
                    return {
                        'success': False,
                        'error': 'No connection configuration stored for this data source. Please re-save the connection (Data sources → Edit → Save).'
                    }

                # CRITICAL: Decrypt credentials before using for schema fetch (use a copy so we do not mutate stored shape)
                try:
                    from src.modules.data.utils.credentials import decrypt_credentials
                    config = decrypt_credentials(dict(config))
                    logger.debug("Decrypted credentials for schema fetch: %s", data_source_id)
                except Exception as decrypt_error:
                    logger.warning("Could not decrypt credentials for schema fetch (may not be encrypted): %s", decrypt_error)

                # Ensure type is set for connector (stored config might use db_type on the row)
                if not config.get('type') and getattr(data_source, 'db_type', None):
                    config['type'] = data_source.db_type

                # SQL Server (and other DBs) require username/password from the stored config — no defaults
                db_type = (config.get('type') or getattr(data_source, 'db_type', '') or '').lower()
                if db_type == 'sqlserver':
                    if not (config.get('username') or str(config.get('username', '')).strip()):
                        return {
                            'success': False,
                            'error': 'Stored connection has no username. Please re-save this data source with the correct SQL Server username (Data sources → Edit → Save).'
                        }
                    if not (config.get('password') is not None and str(config.get('password', '')).strip()):
                        return {
                            'success': False,
                            'error': 'Stored connection has no password. Please re-save this data source with the correct SQL Server password (Data sources → Edit → Save).'
                        }

                # Return in-memory cached schema if still fresh
                import time as _time
                _cached = self._schema_cache.get(data_source_id)
                if not force_refresh and _cached:
                    _cached_schema, _cached_at = _cached
                    if _time.monotonic() - _cached_at < self._schema_cache_ttl:
                        logger.info("✅ Returning cached schema for %s", data_source_id)
                        return {
                            'success': True,
                            'schema': _cached_schema,
                            'data_source': {
                                'id': data_source.id,
                                'name': data_source.name,
                                'type': data_source.type,
                                'db_type': data_source.db_type,
                                'row_count': data_source.row_count,
                            },
                        }

                # Return stored schema immediately if it has tables and no forced refresh
                if not force_refresh and schema_info and isinstance(schema_info, dict) and schema_info.get('tables'):
                    logger.info("✅ Returning stored schema for %s (%d tables)", data_source_id, len(schema_info['tables']))
                    self._schema_cache[data_source_id] = (schema_info, _time.monotonic())
                    return {
                        'success': True,
                        'schema': schema_info,
                        'data_source': {
                            'id': data_source.id,
                            'name': data_source.name,
                            'type': data_source.type,
                            'db_type': data_source.db_type,
                            'row_count': data_source.row_count,
                        },
                    }

                # Try to get live schema from the database
                try:
                    logger.info(f"🔍 Fetching live schema for {data_source_id}, db_type: {data_source.db_type}, type: {data_source.type}")
                    live_schema = await asyncio.wait_for(
                        self._fetch_live_database_schema(config),
                        timeout=25.0,
                    )
                    logger.info(f"📊 Live schema result: success={live_schema.get('success')}, tables_count={len(live_schema.get('tables', []))}, schemas_count={len(live_schema.get('schemas', []))}")
                    
                    if live_schema['success']:
                        tables = live_schema.get('tables', [])
                        schemas = live_schema.get('schemas', [])
                        
                        # Update the stored schema with live data
                        updated_schema = {
                            **schema_info,
                            'tables': tables,
                            'schemas': schemas,
                            'last_updated': datetime.now().isoformat()
                        }
                        
                        logger.info(f"✅ Schema fetched successfully: {len(tables)} tables, {len(schemas)} schemas")

                        # Cache in memory
                        self._schema_cache[data_source_id] = (updated_schema, _time.monotonic())

                        # Update the database record
                        data_source.schema = json.dumps(updated_schema)
                        data_source.row_count = live_schema.get('total_rows', 0)
                        data_source.updated_at = datetime.now()
                        await db.commit()
                        # Build schema index for Schema RAG (NL2SQL table ranking) on schema refresh.
                        try:
                            from src.modules.ai.services.schema_index_service import build_schema_index_for_data_source
                            build_result = await build_schema_index_for_data_source(db, str(data_source.id), updated_schema)
                            await db.commit()
                            logger.info(
                                "Schema index built on refresh: data_source_id=%s, tables_indexed=%s",
                                data_source.id,
                                build_result.get("tables_indexed", 0),
                            )
                        except Exception as idx_err:
                            logger.warning("Schema index build after schema refresh failed: %s", idx_err)
                        
                        return {
                            'success': True,
                            'schema': updated_schema,
                            'data_source': {
                                'id': data_source.id,
                                'name': data_source.name,
                                'type': data_source.type,
                                'db_type': data_source.db_type,
                                'row_count': data_source.row_count
                            }
                        }
                    else:
                        # Live fetch failed (e.g. wrong credentials) — return error so UI can show it
                        return {
                            'success': False,
                            'error': live_schema.get('error') or 'Schema fetch failed',
                            'schema': schema_info,
                            'data_source': {
                                'id': data_source.id,
                                'name': data_source.name,
                                'type': data_source.type,
                                'db_type': data_source.db_type,
                                'row_count': data_source.row_count
                            }
                        }

                except asyncio.TimeoutError:
                    logger.warning("Live schema fetch timed out after 25s for %s", data_source_id)
                    return {
                        'success': False,
                        'error': 'Schema fetch timed out. The database may be slow or unreachable.',
                        'schema': schema_info,
                        'data_source': {
                            'id': data_source.id,
                            'name': data_source.name,
                            'type': data_source.type,
                            'db_type': data_source.db_type,
                            'row_count': data_source.row_count
                        }
                    }
                except Exception as live_error:
                    logger.warning("Live schema fetch failed: %s", str(live_error))
                    return {
                        'success': False,
                        'error': str(live_error),
                        'schema': schema_info,
                        'data_source': {
                            'id': data_source.id,
                            'name': data_source.name,
                            'type': data_source.type,
                            'db_type': data_source.db_type,
                            'row_count': data_source.row_count
                        }
                    }
                    
        except Exception as e:
            logger.error(f"❌ Failed to get database schema: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    async def _fetch_live_database_schema(self, config: Dict[str, Any]) -> Dict[str, Any]:
        """Fetch live schema from the database connection"""
        try:
            # CRITICAL: Ensure credentials are decrypted before using
            try:
                from src.modules.data.utils.credentials import decrypt_credentials
                config = decrypt_credentials(config)
                logger.debug(f"✅ Decrypted credentials for schema fetch")
            except Exception as decrypt_error:
                logger.debug(f"Credentials may not be encrypted: {decrypt_error}")
            
            db_type = config.get('type', '').lower()

            # NoSQL: fetch schema via dedicated connectors
            if db_type in NOSQL_TYPES:
                return await self._fetch_nosql_schema(db_type, config)

            # Get supported databases from DatabaseConnectorService
            supported_dbs = self.database_connector.get_supported_databases()

            if db_type not in supported_dbs:
                return {
                    'success': False,
                    'error': f'Unsupported database type: {db_type}. Supported: {supported_dbs}'
                }

            # Use DatabaseConnectorService to get schema
            try:
                schema_result = await self.database_connector.get_schema(config)
                if schema_result['success']:
                    return {
                        'success': True,
                        'tables': schema_result.get('tables', []),
                        'schemas': schema_result.get('schemas', []),
                        'total_rows': schema_result.get('total_rows', 0)
                    }
                else:
                    err_msg = schema_result.get('error') or ''
                    logger.warning("⚠️ Schema fetch failed: %s", err_msg)
                    # Do not return fallback for login/credential failures — surface error so API can return 400
                    err_lower = err_msg.lower() if isinstance(err_msg, str) else ''
                    if (
                        '28000' in str(err_msg) or 'login failed' in err_lower
                        or 'authentication' in err_lower or 'auth_failed' in err_lower
                        or 'password is incorrect' in err_lower or '403' in str(err_msg)
                    ):
                        return {
                            'success': False,
                            'error': (
                                'Database login failed with the credentials stored for this data source. '
                                'Open Data sources → select this source → Edit, re-enter the correct username and password, then Save. '
                                'If ENCRYPTION_KEY was not set when you first saved, set it in .env and re-save the connection.'
                            )
                        }
                    return await self._get_enhanced_fallback_schema(config)

            except Exception as schema_error:
                err_str = str(schema_error)
                logger.warning("Schema fetch failed: %s", err_str)
                # If login/credential failure, return a clear message so user re-saves the connection
                if (
                    '28000' in err_str or 'Login failed' in err_str
                    or 'authentication' in err_str.lower()
                    or 'AUTHENTICATION_FAILED' in err_str or '403' in err_str
                    or 'password is incorrect' in err_str
                ):
                    return {
                        'success': False,
                        'error': (
                            'Database login failed with the credentials stored for this data source. '
                            'Open Data sources → select this source → Edit, re-enter the correct username and password, then Save. '
                            'If ENCRYPTION_KEY was not set when you first saved, set it in .env and re-save the connection.'
                        )
                    }
                return await self._get_enhanced_fallback_schema(config)

        except Exception as e:
            logger.error("Live schema fetch failed: %s", str(e))
            return {
                'success': False,
                'error': str(e)
            }

    async def _get_enhanced_fallback_schema(self, config: Dict[str, Any]) -> Dict[str, Any]:
        """Get fallback schema - only returns empty schema, no mock data"""
        try:
            db_type = config.get('type', '').lower()
            db_name = config.get('database', 'unknown')
            
            # CRITICAL: Do not return mock data - only return empty schema
            # Real schema should be fetched via get_database_schema which uses SQLAlchemy
            logger.warning(f"⚠️ Fallback schema called for {db_type} - returning empty schema. Real schema should be fetched via database connection.")
            
            # Return failure — callers must not treat an unreachable DB as an empty schema.
            # Returning success:True here caused get_database_schema to overwrite the stored
            # schema with empty tables, permanently destroying it on every failed connection.
            return {
                'success': False,
                'error': 'Schema could not be fetched. Please ensure the database connection is valid and try again.',
                'tables': [],
                'schemas': [],
                'total_rows': 0,
            }
            
        except Exception as e:
            logger.error(f"❌ Enhanced fallback schema failed: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }

    # Utility methods
    def _get_file_extension(self, filename: str) -> str:
        """Get file extension"""
        return Path(filename).suffix.lower().lstrip('.')

    # Database connector methods
    async def _create_database_connector(self, config: Dict[str, Any]):
        """Create database connector"""
        # This is handled by the database_connector service
        return await self.database_connector.create_connection(config)

    def _convert_to_cube_query(self, query: Dict[str, Any], cube_data_source: Dict[str, Any]) -> Dict[str, Any]:
        """Convert our query format to Cube.js query format"""
        cube_query = {}
        
        # Add measures (default to count if none specified)
        cube_query['measures'] = query.get('measures', ['count'])
        
        # Add dimensions
        if query.get('dimensions'):
            cube_query['dimensions'] = query['dimensions']
        
        # Add time dimensions
        if query.get('time_dimensions'):
            cube_query['timeDimensions'] = query['time_dimensions']
        
        # Add filters
        if query.get('filters'):
            cube_query['filters'] = self._convert_filters_to_cube_format(query['filters'])
        
        # Add sorting
        if query.get('sort'):
            sort_config = query['sort']
            cube_query['order'] = [[sort_config['column'], sort_config.get('direction', 'asc')]]
        
        # Add limit
        if query.get('limit'):
            cube_query['limit'] = query['limit']
        
        return cube_query

    def _convert_filters_to_cube_format(self, filters: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Convert filters to Cube.js format"""
        cube_filters = []
        
        for filter_item in filters:
            cube_filter = {
                'member': filter_item['column'],
                'operator': self._map_operator_to_cube(filter_item['operator']),
                'values': [filter_item['value']]
            }
            cube_filters.append(cube_filter)
        
        return cube_filters

    def _map_operator_to_cube(self, operator: str) -> str:
        """Map our operators to Cube.js operators"""
        operator_mapping = {
            'equals': 'equals',
            'contains': 'contains',
            'greater_than': 'gt',
            'less_than': 'lt',
            'greater_equal': 'gte',
            'less_equal': 'lte'
        }
        return operator_mapping.get(operator, 'equals')

    async def get_supported_databases(self) -> Dict[str, Any]:
        """Get supported database types"""
        supported_dbs = self.database_connector.get_supported_databases()
        return {
            'success': True,
            'supported_databases': [
                {
                    'type': db_type,
                    'name': db_type.title(),
                    'driver': 'sqlalchemy',
                }
                for db_type in supported_dbs
            ]
        }

    def _auto_detect_delimiter(self, file_path: str) -> str:
        """Auto-detect CSV delimiter"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                first_line = f.readline().strip()
            
            if not first_line:
                return ','
            
            # Common delimiters to test
            delimiters = [',', ';', '\t', '|', ' ']
            max_fields = 0
            best_delimiter = ','
            
            for delimiter in delimiters:
                fields = first_line.split(delimiter)
                if len(fields) > max_fields:
                    max_fields = len(fields)
                    best_delimiter = delimiter
            
            return best_delimiter
            
        except Exception:
            return ','

    def _auto_detect_encoding(self, file_path: str) -> str:
        """Auto-detect file encoding"""
        try:
            import chardet
            
            with open(file_path, 'rb') as f:
                raw_data = f.read(10000)  # Read first 10KB
            
            result = chardet.detect(raw_data)
            encoding = result['encoding']
            confidence = result['confidence']
            
            if confidence > 0.7:
                return encoding
            else:
                return 'utf-8'
                
        except ImportError:
            # chardet not available, try common encodings
            encodings = ['utf-8', 'latin-1', 'cp1252']
            
            for encoding in encodings:
                try:
                    with open(file_path, 'r', encoding=encoding) as f:
                        f.readline()
                    return encoding
                except UnicodeDecodeError:
                    continue
            
            return 'utf-8'
        except Exception:
            return 'utf-8'

    async def get_file_preview(self, file_path: str, file_format: str, options: dict = None) -> dict:
        """Get enhanced file preview with auto-detection"""
        try:
            if options is None:
                options = {}
            
            preview_data = {
                'format': file_format,
                'size': os.path.getsize(file_path),
                'auto_detected_config': {},
                'sample_data': [],
                'schema_preview': {},
                'processing_options': {}
            }
            
            if file_format in ['csv', 'tsv']:
                # Auto-detect delimiter and encoding
                delimiter = options.get('delimiter', self._auto_detect_delimiter(file_path))
                encoding = options.get('encoding', self._auto_detect_encoding(file_path))
                
                preview_data['auto_detected_config'] = {
                    'delimiter': delimiter,
                    'encoding': encoding
                }
                
                # Read first few lines for preview
                df_preview = pd.read_csv(file_path, delimiter=delimiter, encoding=encoding, nrows=PREVIEW_ROWS)
                preview_data['sample_data'] = df_preview.to_dict('records')
                preview_data['schema_preview'] = {
                    'columns': list(df_preview.columns),
                    'data_types': df_preview.dtypes.to_dict(),
                    'row_count_preview': len(df_preview)
                }
                
                # Get processing options
                preview_data['processing_options'] = {
                    'delimiters': self.file_processing_configs['csv']['delimiters'],
                    'encodings': self.file_processing_configs['csv']['encodings']
                }
                
            elif file_format in ['xlsx', 'xls']:
                # Get sheet information
                import openpyxl
                workbook = openpyxl.load_workbook(file_path, read_only=True)
                sheets = workbook.sheetnames
                
                preview_data['auto_detected_config'] = {
                    'sheets': sheets,
                    'default_sheet': sheets[0] if sheets else None
                }
                
                # Read first sheet for preview
                df_preview = pd.read_excel(file_path, sheet_name=sheets[0], nrows=PREVIEW_ROWS)
                preview_data['sample_data'] = df_preview.to_dict('records')
                preview_data['schema_preview'] = {
                    'columns': list(df_preview.columns),
                    'data_types': df_preview.dtypes.to_dict(),
                    'row_count_preview': len(df_preview)
                }
                
                preview_data['processing_options'] = {
                    'sheets': sheets,
                    'header_row_options': [0, 1, 2],
                    'skip_rows_options': [0, 1, 2, 3]
                }
                
            elif file_format == 'parquet':
                # Get parquet metadata
                import pyarrow.parquet as pq
                parquet_file = pq.ParquetFile(file_path)
                metadata = parquet_file.metadata
                
                preview_data['auto_detected_config'] = {
                    'columns': list(metadata.schema.names),
                    'row_groups': metadata.num_row_groups,
                    'total_rows': metadata.num_rows
                }
                
                # Read sample data
                df_preview = parquet_file.read().to_pandas().head(PREVIEW_ROWS)
                preview_data['sample_data'] = df_preview.to_dict('records')
                preview_data['schema_preview'] = {
                    'columns': list(df_preview.columns),
                    'data_types': df_preview.dtypes.to_dict(),
                    'row_count_preview': len(df_preview)
                }
                
            elif file_format == 'json':
                # Get JSON structure info
                with open(file_path, 'r', encoding='utf-8') as f:
                    json_data = json.load(f)
                
                if isinstance(json_data, list):
                    preview_data['auto_detected_config'] = {
                        'structure': 'array',
                        'item_count': len(json_data)
                    }
                    sample_data = json_data[:PREVIEW_ROWS] if len(json_data) > PREVIEW_ROWS else json_data
                else:
                    preview_data['auto_detected_config'] = {
                        'structure': 'object',
                        'keys': list(json_data.keys())
                    }
                    sample_data = [json_data]
                
                preview_data['sample_data'] = sample_data
                preview_data['schema_preview'] = {
                    'structure_type': preview_data['auto_detected_config']['structure'],
                    'sample_keys': list(sample_data[0].keys()) if sample_data else []
                }
            
            return {
                'success': True,
                'preview': preview_data
            }
            
        except Exception as e:
            logger.error(f"❌ File preview generation failed: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def _initialize_demo_data(self):
        """Initialize demo data sources for testing and demonstration"""
        try:
            logger.info("🚀 Initializing demo data sources...")
            
            # Demo sales data
            demo_sales = {
                'id': 'demo_sales_data',
                'name': 'Demo Sales Data',
                'type': 'file',
                'format': 'csv',
                'description': 'Sample sales data for demonstration and testing',
                'created_at': datetime.now().isoformat(),
                'updated_at': datetime.now().isoformat(),
                'size': 24576,  # 24KB
                'row_count': 1000,
                'schema': {
                    'columns': [
                        {'name': 'date', 'type': 'datetime', 'nullable': False},
                        {'name': 'product_name', 'type': 'string', 'nullable': False},
                        {'name': 'category', 'type': 'string', 'nullable': False},
                        {'name': 'sales_amount', 'type': 'numeric', 'nullable': False},
                        {'name': 'quantity', 'type': 'integer', 'nullable': False},
                        {'name': 'region', 'type': 'string', 'nullable': False},
                        {'name': 'customer_id', 'type': 'string', 'nullable': False}
                    ]
                },
                'sample_data': [
                    {'date': '2024-01-15', 'product_name': 'Laptop Pro', 'category': 'Electronics', 'sales_amount': 1299.99, 'quantity': 1, 'region': 'North', 'customer_id': 'C001'},
                    {'date': '2024-01-15', 'product_name': 'Wireless Mouse', 'category': 'Accessories', 'sales_amount': 29.99, 'quantity': 2, 'region': 'South', 'customer_id': 'C002'},
                    {'date': '2024-01-16', 'product_name': 'Monitor 4K', 'category': 'Electronics', 'sales_amount': 599.99, 'quantity': 1, 'region': 'East', 'customer_id': 'C003'},
                    {'date': '2024-01-16', 'product_name': 'Keyboard', 'category': 'Accessories', 'sales_amount': 89.99, 'quantity': 1, 'region': 'West', 'customer_id': 'C004'},
                    {'date': '2024-01-17', 'product_name': 'Tablet', 'category': 'Electronics', 'sales_amount': 399.99, 'quantity': 1, 'region': 'North', 'customer_id': 'C005'}
                ],
                'metadata': {
                    'source': 'demo_data',
                    'business_domain': 'retail',
                    'data_quality': 'high',
                    'last_updated': datetime.now().isoformat()
                }
            }
            
            # Demo customer data
            demo_customers = {
                'id': 'demo_customers_data',
                'name': 'Demo Customer Data',
                'type': 'file',
                'format': 'csv',
                'description': 'Sample customer data for demonstration and testing',
                'created_at': datetime.now().isoformat(),
                'updated_at': datetime.now().isoformat(),
                'size': 15360,  # 15KB
                'row_count': 500,
                'schema': {
                    'columns': [
                        {'name': 'customer_id', 'type': 'string', 'nullable': False},
                        {'name': 'first_name', 'type': 'string', 'nullable': False},
                        {'name': 'last_name', 'type': 'string', 'nullable': False},
                        {'name': 'email', 'type': 'string', 'nullable': False},
                        {'name': 'age', 'type': 'integer', 'nullable': True},
                        {'name': 'city', 'type': 'string', 'nullable': False},
                        {'name': 'country', 'type': 'string', 'nullable': False},
                        {'name': 'registration_date', 'type': 'datetime', 'nullable': False}
                    ]
                },
                'sample_data': [
                    {'customer_id': 'C001', 'first_name': 'John', 'last_name': 'Doe', 'email': 'john.doe@email.com', 'age': 35, 'city': 'New York', 'country': 'USA', 'registration_date': '2023-01-15'},
                    {'customer_id': 'C002', 'first_name': 'Jane', 'last_name': 'Smith', 'email': 'jane.smith@email.com', 'age': 28, 'city': 'Los Angeles', 'country': 'USA', 'registration_date': '2023-02-20'},
                    {'customer_id': 'C003', 'first_name': 'Bob', 'last_name': 'Johnson', 'email': 'bob.johnson@email.com', 'age': 42, 'city': 'Chicago', 'country': 'USA', 'registration_date': '2023-03-10'},
                    {'customer_id': 'C004', 'first_name': 'Alice', 'last_name': 'Brown', 'email': 'alice.brown@email.com', 'age': 31, 'city': 'Houston', 'country': 'USA', 'registration_date': '2023-04-05'},
                    {'customer_id': 'C005', 'first_name': 'Charlie', 'last_name': 'Wilson', 'email': 'charlie.wilson@email.com', 'age': 39, 'city': 'Phoenix', 'country': 'USA', 'registration_date': '2023-05-12'}
                ],
                'metadata': {
                    'source': 'demo_data',
                    'business_domain': 'customer_management',
                    'data_quality': 'high',
                    'last_updated': datetime.now().isoformat()
                }
            }
            
            # Add demo data sources to the registry
            self.data_sources['demo_sales_data'] = demo_sales
            self.data_sources['demo_customers_data'] = demo_customers

            # Also expose convenient aliases commonly used by the UI
            self.data_sources['duckdb_local'] = self.data_sources['demo_sales_data']
            self.data_sources['csv_sales'] = self.data_sources['demo_sales_data']
            self.data_sources['snowflake_warehouse'] = self.data_sources['demo_customers_data']
            self.data_sources['postgresql_prod'] = self.data_sources['demo_customers_data']
            
            logger.info(f"✅ Demo data sources initialized successfully: {list(self.data_sources.keys())}")
            logger.info(f"✅ Total demo sources: {len(self.data_sources)}")
            
        except Exception as e:
            logger.error(f"❌ Failed to initialize demo data: {str(e)}")
            import traceback
            logger.error(f"❌ Traceback: {traceback.format_exc()}")
    
    def _get_current_user_id(self) -> int:
        """Unused legacy helper — API handlers pass ``user_id`` from JWT. Do not use for authorization."""
        logger.debug("_get_current_user_id is deprecated; pass user_id from request dependencies")
        return 0

    # 🏗️ PROJECT-SCOPED DATA SOURCE METHODS

    async def get_project_data_sources(
        self, 
        organization_id: str, 
        project_id: str, 
        user_id: str = None, 
        offset: int = 0, 
        limit: int = None
    ) -> List[Dict[str, Any]]:
        """Get data sources for a specific project (project-scoped)"""
        if limit is None:
            limit = DEFAULT_LIST_PAGE_LIMIT
        try:
            logger.info(f"🔍 Getting project data sources for project {project_id} in organization {organization_id}")
            
            from src.modules.data.models import DataSource
            from src.modules.data.models import ProjectDataSource
            from src.db.session import async_session

            async with async_session() as db:
                from sqlalchemy import select, join

                # Join data sources with project data sources
                query = (
                    select(DataSource)
                    .join(ProjectDataSource, DataSource.id == ProjectDataSource.data_source_id)
                    .where(
                        # Accept numeric IDs or slugs; if not int, skip casting
                        ProjectDataSource.project_id == (int(project_id) if str(project_id).isdigit() else ProjectDataSource.project_id),
                        DataSource.is_active == True,
                        ProjectDataSource.is_active == True
                    )
                    .offset(offset)
                    .limit(limit)
                )
                
                result = await db.execute(query)
                data_sources = result.scalars().all()
                
                return [
                    {
                        'id': ds.id,
                        'name': ds.name,
                        'type': ds.type,
                        'format': ds.format,
                        'size': ds.size,
                        'row_count': ds.row_count,
                        'schema': ds.schema,
                        'organization_id': organization_id,
                        'project_id': project_id,
                        'created_at': ds.created_at.isoformat() if ds.created_at else None,
                        'updated_at': ds.updated_at.isoformat() if ds.updated_at else None,
                        'last_accessed': ds.last_accessed.isoformat() if ds.last_accessed else None
                    }
                    for ds in data_sources
                ]
        except Exception as e:
            logger.error(f"❌ Failed to get project data sources: {str(e)}")
            return []

    async def create_project_data_source(
        self, 
        organization_id: str, 
        project_id: str, 
        data_source_data: Dict[str, Any],
        user_id: Optional[str] = None
    ) -> Dict[str, Any]:
        """Create a new data source for a specific project
        
        NOTE: This method is legacy and may be deprecated. user_id should be provided.
        """
        # Extract user_id from data_source_data if not provided directly
        if not user_id:
            user_id = data_source_data.get('user_id')
        
        if not user_id:
            raise ValueError("user_id is required for data source creation. Provide it in data_source_data or as a parameter.")
        
        try:
            logger.info(f"📊 Creating project data source for project {project_id} (user_id={user_id})")
            
            from src.modules.data.models import DataSource
            from src.modules.data.models import ProjectDataSource
            from src.db.session import async_session

            # Generate a stable unique id for this create call so we can
            # expose a preview entry immediately in the in-memory registry
            import uuid
            generated_id = f"ds_{uuid.uuid4().hex}"

            # Store a preview entry so immediate calls (modeling) can read sample data
            try:
                self.data_sources[generated_id] = {
                    'id': generated_id,
                    'name': data_source_data.get('name'),
                    'type': data_source_data.get('type'),
                    'format': data_source_data.get('format'),
                    'schema': data_source_data.get('schema'),
                    'data': data_source_data.get('data') or data_source_data.get('sample_data') or [],
                    'created_at': datetime.now().isoformat(),
                    'updated_at': datetime.now().isoformat(),
                    'organization_id': organization_id,
                    'project_id': project_id,
                    'is_active': True,
                }
            except Exception:
                pass

            async def _do_create():
                async with async_session() as db:
                    # Create the data source
                    # Build metadata and connection config safely
                    metadata = data_source_data.get('metadata', {}) or {}
                    # carry over description into metadata to avoid unknown kwargs on model
                    if data_source_data.get('description'):
                        metadata.setdefault('description', data_source_data.get('description'))

                    connection_config = data_source_data.get('config') or {}
                    # Decrypt any encrypted credentials before using
                    try:
                        from src.modules.data.utils.credentials import decrypt_credentials

                        connection_config = decrypt_credentials(connection_config)
                    except Exception:
                        pass
                    db_type = connection_config.get('type') or data_source_data.get('db_type')

                    # Prepare data source values so we can use them in both
                    # async and sync (test) paths.
                    data_source_values = {
                        'id': generated_id,
                        'name': data_source_data['name'],
                        'type': data_source_data['type'],
                        'format': data_source_data.get('format'),
                        'db_type': db_type,
                        'schema': data_source_data.get('schema'),
                        'connection_config': connection_config,
                    # Allow callers to provide inline sample rows for file sources
                        'sample_data': data_source_data.get('data') or data_source_data.get('sample_data') or [],
                        # Use actual user_id (required for data isolation)
                        'user_id': str(user_id),
                        'is_active': True,
                        'created_at': datetime.now(),
                        'updated_at': datetime.now(),
                    }

                    # create SQLAlchemy DataSource instance for async path
                    # If the generated id already exists due to previous attempts, reuse it
                    try:
                        from src.db.session import get_sync_engine
                        eng_check = get_sync_engine()
                        import sqlalchemy as sa
                        with eng_check.connect() as conn:
                            row = conn.execute(sa.text("SELECT id FROM data_sources WHERE id = :id"), {"id": generated_id}).fetchone()
                            if row and row[0]:
                                # already exists; return early using existing id
                                return {
                                    'success': True,
                                    'data_source': {
                                        'id': str(row[0]),
                                        'name': data_source_values.get('name'),
                                        'type': data_source_values.get('type'),
                                        'organization_id': organization_id,
                                        'project_id': project_id,
                                    },
                                    'data_source_id': str(row[0]),
                                    'id': str(row[0]),
                                    'message': 'Data source already exists'
                                }
                    except Exception:
                        # non-fatal; continue to create
                        pass

                    data_source = DataSource(**data_source_values)

                    # Persist using the global database helper which creates its
                    # own session per call and serializes operations.
                    from src.db.session import get_db
                    from sqlalchemy.exc import IntegrityError as SAIntegrityError

                    db_helper = get_db()
                    # Save data_source via helper (creates new session internally)
                    # Retry on unique key collisions by regenerating id
                    max_attempts = 3
                    attempt = 0
                    while True:
                        try:
                            attempt += 1
                            await db_helper.add(data_source)
                            break
                        except SAIntegrityError as sie:
                            if attempt >= max_attempts:
                                raise
                            # regenerate id and retry
                            import uuid as _uuid
                            new_id = f"ds_{_uuid.uuid4().hex}"
                            data_source_values['id'] = new_id
                            # recreate SQLAlchemy instance with new id
                            data_source = DataSource(**data_source_values)
                            continue
                    # Persist sample_data to the DB if provided
                    try:
                        if data_source_values.get('sample_data'):
                            upd = sa.text("UPDATE data_sources SET sample_data = :sd WHERE id = :id")
                            with get_sync_engine().begin() as sync_conn:
                                sync_conn.execute(upd, {"sd": json.dumps(data_source_values.get('sample_data')), "id": data_source.id})
                    except Exception:
                        # Non-fatal; sample persistence best-effort
                        logger.debug('Failed to persist sample_data for %s', data_source.id)

                    # Link to project and persist
                    project_data_source = ProjectDataSource(
                        project_id=int(project_id),
                        data_source_id=data_source.id,
                        data_source_type=data_source_data['type'],
                        is_active=True,
                        added_at=datetime.now(),
                    )

                    await db_helper.add(project_data_source)

                    return {
                        'success': True,
                        'data_source': {
                            'id': data_source.id,
                            'name': data_source.name,
                            'type': data_source.type,
                            'organization_id': organization_id,
                            'project_id': project_id,
                        },
                        'data_source_id': data_source.id,
                        'id': data_source.id,
                        'message': 'Data source created successfully',
                    }

            # Use global async write queue to serialize DB writes safely.
            # Execute the create via a sync worker (thread) using the sync engine
            # to avoid asyncpg 'another operation is in progress' errors when
            # other async operations are running on the same event loop.
            try:
                from src.db.write_queue import write_queue

                # Prepare a sync callable that mirrors the async create path
                def _do_create_sync():
                    try:
                        # Rebuild connection/config values
                        connection_config = data_source_data.get('config') or {}
                        try:
                            from src.modules.data.utils.credentials import decrypt_credentials

                            connection_config = decrypt_credentials(connection_config)
                        except Exception:
                            pass

                        db_type = connection_config.get('type') or data_source_data.get('db_type')
                        ds_vals = {
                            'id': generated_id,
                            'name': data_source_data['name'],
                            'type': data_source_data['type'],
                            'format': data_source_data.get('format'),
                            'db_type': db_type,
                            'schema': data_source_data.get('schema'),
                            'connection_config': connection_config,
                            'sample_data': data_source_data.get('data') or data_source_data.get('sample_data') or [],
                            'user_id': str(organization_id),
                            'is_active': True,
                            'created_at': datetime.now(),
                            'updated_at': datetime.now()
                        }

                        from src.db.session import get_sync_engine
                        eng = get_sync_engine()
                        with eng.begin() as conn:
                            insert_ds = sa.text(
                                "INSERT INTO data_sources (id, name, type, format, db_type, size, row_count, schema, description, connection_config, file_path, original_filename, created_at, updated_at, project_id, is_active, last_accessed) "
                                "VALUES (:id, :name, :type, :format, :db_type, :size, :row_count, :schema, :description, :connection_config, :file_path, :original_filename, :created_at, :updated_at, :project_id, :is_active, :last_accessed) "
                                "ON CONFLICT (id) DO NOTHING"
                            )
                            # Convert project_id to UUID if it's a string
                            from uuid import UUID
                            project_id_val = ds_vals.get('project_id')
                            if project_id_val and isinstance(project_id_val, str):
                                try:
                                    project_id_val = UUID(project_id_val)
                                except ValueError:
                                    logger.warning(f"Invalid project_id format: {project_id_val}")
                                    project_id_val = None
                            
                            params = {
                                'id': ds_vals.get('id'),
                                'name': ds_vals.get('name'),
                                'type': ds_vals.get('type'),
                                'format': ds_vals.get('format'),
                                'db_type': ds_vals.get('db_type'),
                                'size': ds_vals.get('size'),
                                'row_count': ds_vals.get('row_count'),
                                'schema': json.dumps(ds_vals.get('schema')) if ds_vals.get('schema') is not None else None,
                                'description': ds_vals.get('description') if ds_vals.get('description') is not None else None,
                                'connection_config': json.dumps(ds_vals.get('connection_config')) if ds_vals.get('connection_config') is not None else None,
                                'file_path': ds_vals.get('file_path') if ds_vals.get('file_path') is not None else None,
                                'original_filename': ds_vals.get('original_filename') if ds_vals.get('original_filename') is not None else None,
                                'created_at': ds_vals.get('created_at'),
                                'updated_at': ds_vals.get('updated_at'),
                                'project_id': project_id_val,
                                'is_active': ds_vals.get('is_active'),
                                'last_accessed': ds_vals.get('last_accessed')
                            }
                            # Retry insert on unique violation
                            max_sync_attempts = 3
                            sync_attempt = 0
                            while True:
                                try:
                                    conn.execute(insert_ds, params)

                                    # ensure id exists (select back)
                                    sel = sa.text("SELECT id FROM data_sources WHERE id = :id LIMIT 1")
                                    rsel = conn.execute(sel, {"id": params['id']})
                                    row = rsel.fetchone()
                                    if not row:
                                        # nothing inserted (conflict); try to find by name+project
                                        sel2 = sa.text(
                                            "SELECT ds.id FROM data_sources ds JOIN project_data_source pds ON pds.data_source_id = ds.id "
                                            "WHERE ds.name = :name AND pds.project_id = :pid LIMIT 1"
                                        )
                                        r2 = conn.execute(sel2, {"name": ds_vals.get('name'), "pid": int(project_id)})
                                        rr = r2.fetchone()
                                        if rr and rr[0]:
                                            params['id'] = rr[0]
                                            break
                                    else:
                                        # inserted successfully
                                        break
                                except Exception as e:
                                    # detect unique violation (asyncpg / psycopg2 messages vary)
                                    msg = str(e).lower()
                                    if 'unique' in msg or 'duplicate' in msg:
                                        sync_attempt += 1
                                        if sync_attempt >= max_sync_attempts:
                                            raise
                                        import uuid as _uuid
                                        new_id = f"ds_{_uuid.uuid4().hex}"
                                        params['id'] = new_id
                                        ds_vals['id'] = new_id
                                        # loop and retry with new id
                                        continue
                                    raise
                            # Persist sample_data if present
                            if ds_vals.get('sample_data'):
                                try:
                                    upd = sa.text("UPDATE data_sources SET sample_data = :sd WHERE id = :id")
                                    conn.execute(upd, {"sd": json.dumps(ds_vals.get('sample_data')), "id": ds_vals.get('id')})
                                except Exception:
                                    logger.debug('Failed to persist sample_data in sync path for %s', ds_vals.get('id'))
                            insert_link = sa.text(
                                "INSERT INTO project_data_source (id, project_id, data_source_id, data_source_type, is_active, added_at) VALUES (gen_random_uuid(), :project_id, :data_source_id, :data_source_type, :is_active, :added_at)"
                            )
                            conn.execute(insert_link, {"project_id": int(project_id), "data_source_id": ds_vals.get('id'), "data_source_type": data_source_data['type'], "is_active": True, "added_at": datetime.now()})

                        # Update in-memory registry
                        # Persist a copy in the in-memory registry including sample rows if provided
                        self.data_sources[ds_vals.get('id')] = {
                            'id': ds_vals.get('id'),
                            'name': ds_vals.get('name'),
                            'type': ds_vals.get('type'),
                            'format': ds_vals.get('format'),
                            'schema': ds_vals.get('schema'),
                            'data': ds_vals.get('data') or [],
                            'created_at': ds_vals.get('created_at').isoformat() if ds_vals.get('created_at') else None,
                            'updated_at': ds_vals.get('updated_at').isoformat() if ds_vals.get('updated_at') else None,
                            'is_active': True,
                            'organization_id': organization_id,
                            'project_id': project_id
                        }
                        return {
                            'success': True,
                            'data_source': {
                                'id': ds_vals.get('id'),
                                'name': ds_vals.get('name'),
                                'type': ds_vals.get('type'),
                                'organization_id': organization_id,
                                'project_id': project_id
                            },
                            'data_source_id': ds_vals.get('id'),
                            'id': ds_vals.get('id'),
                            'message': 'Data source created (queued sync)'
                        }
                    except Exception as e:
                        logger.exception("Sync queued create failed: %s", e)
                        return {'success': False, 'error': str(e)}

                # Serialize DB writes via the write queue to avoid asyncpg "another operation is in progress" errors
                # Use the sync callable so all DB writes run via the sync engine in a worker thread.
                # Retry enqueue a few times in case of transient connection/loop issues, then fall back to async path.
                import asyncio
                max_enqueue_attempts = 3
                for attempt in range(1, max_enqueue_attempts + 1):
                    try:
                        return await write_queue.enqueue(_do_create_sync)
                    except Exception as enqueue_exc:
                        logger.warning("write_queue.enqueue attempt %d/%d failed: %s", attempt, max_enqueue_attempts, enqueue_exc)
                        if attempt == max_enqueue_attempts:
                            logger.exception("write_queue.enqueue failed after %d attempts, falling back to direct async create", max_enqueue_attempts)
                            return await _do_create()
                        await asyncio.sleep(0.4 * attempt)
            except Exception:
                # If queue not available, fall back to direct async create
                return await _do_create()
        except Exception as e:
            # Log full exception with traceback to help debugging in tests
            logger.exception("❌ Failed to create project data source")
            return {
                'success': False,
                'error': repr(e) or str(e) or 'Unknown error'
            }

    async def get_project_data_source(
        self, 
        organization_id: str, 
        project_id: str, 
        data_source_id: str
    ) -> Dict[str, Any]:
        """Get a specific data source for a project"""
        try:
            logger.info(f"🔍 Getting project data source {data_source_id} for project {project_id}")
            
            from src.modules.data.models import DataSource
            from src.modules.data.models import ProjectDataSource
            from src.db.session import async_session

            async with async_session() as db:
                from sqlalchemy import select, join

                query = (
                    select(DataSource)
                    .join(ProjectDataSource, DataSource.id == ProjectDataSource.data_source_id)
                    .where(
                        DataSource.id == data_source_id,
                        ProjectDataSource.project_id == int(project_id),
                        DataSource.is_active == True,
                        ProjectDataSource.is_active == True
                    )
                )
                
                result = await db.execute(query)
                data_source = result.scalar_one_or_none()
                
                if data_source:
                    return {
                        'success': True,
                        'data_source': {
                            'id': data_source.id,
                            'name': data_source.name,
                            'type': data_source.type,
                            'format': data_source.format,
                            'size': data_source.size,
                            'row_count': data_source.row_count,
                            'schema': data_source.schema,
                            'organization_id': organization_id,
                            'project_id': project_id,
                            'created_at': data_source.created_at.isoformat() if data_source.created_at else None,
                            'updated_at': data_source.updated_at.isoformat() if data_source.updated_at else None
                        }
                    }
                else:
                    return {
                        'success': False,
                        'error': 'Data source not found'
                    }
        except Exception as e:
            logger.error(f"❌ Failed to get project data source: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }

    async def execute_project_data_source_query(
        self, 
        organization_id: str, 
        project_id: str, 
        data_source_id: str, 
        query: str
    ) -> Dict[str, Any]:
        """Execute a query on a project data source"""
        try:
            logger.info(f"🔍 Executing project data source query for {data_source_id}")
            
            # First verify the data source belongs to the project
            data_source_result = await self.get_project_data_source(organization_id, project_id, data_source_id)
            if not data_source_result['success']:
                return data_source_result
            
            # Execute the query using the existing method
            return await self.execute_query_on_source(data_source_id, query)
        except Exception as e:
            logger.error(f"❌ Failed to execute project data source query: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }

    async def get_project_data_source_data(
        self, 
        organization_id: str, 
        project_id: str, 
        data_source_id: str, 
        limit: int = None
    ) -> Dict[str, Any]:
        """Get data from a project data source"""
        try:
            logger.info(f"📊 Getting project data source data for {data_source_id}")
            
            # First verify the data source belongs to the project
            data_source_result = await self.get_project_data_source(organization_id, project_id, data_source_id)
            if not data_source_result['success']:
                return data_source_result
            
            # Get the data using the existing method
            return await self.get_data_from_source(data_source_id, limit)
        except Exception as e:
            logger.error(f"❌ Failed to get project data source data: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
